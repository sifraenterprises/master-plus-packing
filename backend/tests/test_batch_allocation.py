"""ASN Manual Batch Allocation - end-to-end tests (TEST/simulation mode).

Covers:
- multi-batch happy path with pause + confirm
- validation errors (over-allocation, unknown batch, wrong total, no-pending 409)
- BATCH-LOW cancel path (no retry loop)
- non-BATCH regression (no pause)
- excel export includes 'Batch Allocations' column
- workflow report drill-down exposes batches on ASN step
"""
import os
import io
import time
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://invoice-master-295.preview.emergentagent.com").rstrip("/")
API = f"{BASE_URL}/api"
PDI_PATH = "/tmp/pdi_test.pdf"


@pytest.fixture(scope="module")
def H():
    r = requests.post(f"{API}/auth/login", json={"username": "admin", "password": "5@Sohangso"})
    assert r.status_code == 200, r.text
    return {"Authorization": f"Bearer {r.json()['token']}"}


def _wait_run_idle(H, timeout=30):
    """Ensure no ASN run is running before we start (avoids 409)."""
    start = time.time()
    while time.time() - start < timeout:
        rs = requests.get(f"{API}/asn/run-status", headers=H).json()
        if not rs.get("running"):
            return rs
        # if awaiting a stale allocation, cancel it
        aw = rs.get("awaiting_allocation")
        if aw:
            requests.post(f"{API}/asn/allocation/cancel", headers=H, json={"record_id": aw["record_id"]})
        time.sleep(1)
    raise AssertionError("previous run never finished")


def _wait_run_done(H, timeout=60):
    start = time.time()
    while time.time() - start < timeout:
        rs = requests.get(f"{API}/asn/run-status", headers=H).json()
        if not rs.get("running"):
            return rs
        time.sleep(1)
    raise AssertionError("run did not finish in time")


def _wait_awaiting(H, timeout=30):
    start = time.time()
    while time.time() - start < timeout:
        rs = requests.get(f"{API}/asn/run-status", headers=H).json()
        if rs.get("awaiting_allocation"):
            return rs["awaiting_allocation"]
        if not rs.get("running"):
            raise AssertionError(f"run finished before awaiting_allocation appeared: {rs}")
        time.sleep(0.5)
    raise AssertionError("awaiting_allocation was never set")


def _seed(H, invoice_no, part_number, qty):
    body = {
        "invoice_number": invoice_no,
        "invoice_date": "2026-01-20",
        "po_number": "5540011947",
        "customer_name": "TAFE",
        "plant": "M04",
        "vehicle_number": "PB10TS1234",
        "transporter_name": "V.R.L. LOGISTICS LIMITED",
        "invoice_total": 118000,
        "gst_total": 18000,
        "cgst": 9000,
        "sgst": 9000,
        "igst": 0,
        "boxes": 2,
        "items": [{"part_number": part_number, "description": "Batch test", "quantity": qty}],
    }
    r = requests.post(f"{API}/master-dispatch", headers=H, json=body)
    assert r.status_code in (200, 201), f"seed md failed: {r.status_code} {r.text}"
    return r.json()


def _import_and_prep(H, invoice_no):
    requests.post(f"{API}/asn/import", headers=H)
    recs = requests.get(f"{API}/asn/records", headers=H, params={"search": invoice_no}).json()["items"]
    assert recs, f"asn record not created for {invoice_no}"
    rec = recs[0]
    rid = rec["id"]
    # edit PO/transporter to ensure non-null
    requests.put(f"{API}/asn/records/{rid}", headers=H, json={
        "po_number": "5540011947", "transporter": "V.R.L. LOGISTICS LIMITED",
        "basic_amount": 100000, "total_amount": 118000,
    })
    with open(PDI_PATH, "rb") as f:
        r = requests.post(f"{API}/asn/records/{rid}/pdi", headers=H,
                          files={"file": ("pdi.pdf", f.read(), "application/pdf")})
    assert r.status_code == 200, r.text
    rec = requests.get(f"{API}/asn/records", headers=H, params={"search": invoice_no}).json()["items"][0]
    assert rec["status"] == "Ready", f"expected Ready got {rec['status']}"
    return rec


def _fetch_record(H, invoice_no):
    return requests.get(f"{API}/asn/records", headers=H, params={"search": invoice_no}).json()["items"][0]


# ---------------- 01: multi-batch happy path ----------------

class TestBatchMultiHappyPath:
    invoice = None
    rid = None
    md_id = None
    asn_number = None

    def test_01a_seed_and_run_multi(self, H):
        _wait_run_idle(H)
        inv = f"TEST/BA-MULTI-{int(time.time())}"
        TestBatchMultiHappyPath.invoice = inv
        md = _seed(H, inv, "BATCH-MULTI-TA", 60)
        TestBatchMultiHappyPath.md_id = md.get("id") or md.get("_id")
        rec = _import_and_prep(H, inv)
        TestBatchMultiHappyPath.rid = rec["id"]
        r = requests.post(f"{API}/asn/run", headers=H, json={"ids": [rec["id"]]})
        assert r.status_code == 200, r.text

    def test_01b_awaiting_shape(self, H):
        aw = _wait_awaiting(H)
        assert aw["record_id"] == TestBatchMultiHappyPath.rid
        assert aw["part_number"] == "BATCH-MULTI-TA"
        assert aw["asn_qty"] == 60
        assert isinstance(aw["batches"], list) and len(aw["batches"]) == 3
        for b in aw["batches"]:
            for k in ("batch_no", "batch_qty", "available_qty"):
                assert k in b, f"missing {k}"
        # Awaiting Allocation status
        rec = _fetch_record(H, TestBatchMultiHappyPath.invoice)
        assert rec["status"] == "Awaiting Allocation", rec["status"]

    def test_01c_validation_total_mismatch(self, H):
        rid = TestBatchMultiHappyPath.rid
        aw = requests.get(f"{API}/asn/run-status", headers=H).json()["awaiting_allocation"]
        batches = aw["batches"]
        # Send total 40 (should be 60) -> 400
        body = {"record_id": rid, "allocations": [
            {"batch_no": batches[0]["batch_no"], "allocate_qty": 10, "consider": True},
            {"batch_no": batches[1]["batch_no"], "allocate_qty": 30, "consider": True},
            {"batch_no": batches[2]["batch_no"], "allocate_qty": 0, "consider": False},
        ]}
        r = requests.post(f"{API}/asn/allocation/confirm", headers=H, json=body)
        assert r.status_code == 400, r.text
        assert "must equal ASN Quantity" in r.text or "must equal" in r.text.lower()

    def test_01d_validation_over_available(self, H):
        rid = TestBatchMultiHappyPath.rid
        aw = requests.get(f"{API}/asn/run-status", headers=H).json()["awaiting_allocation"]
        batches = aw["batches"]
        over = batches[0]["available_qty"] + 100
        body = {"record_id": rid, "allocations": [
            {"batch_no": batches[0]["batch_no"], "allocate_qty": over, "consider": True},
        ]}
        r = requests.post(f"{API}/asn/allocation/confirm", headers=H, json=body)
        assert r.status_code == 400, r.text
        assert "cannot exceed Available Quantity" in r.text

    def test_01e_validation_unknown_batch(self, H):
        rid = TestBatchMultiHappyPath.rid
        body = {"record_id": rid, "allocations": [
            {"batch_no": "NOT-A-REAL-BATCH", "allocate_qty": 60, "consider": True},
        ]}
        r = requests.post(f"{API}/asn/allocation/confirm", headers=H, json=body)
        assert r.status_code == 400, r.text
        assert "Unknown batch" in r.text or "unknown" in r.text.lower()

    def test_01f_confirm_split_success(self, H):
        rid = TestBatchMultiHappyPath.rid
        aw = requests.get(f"{API}/asn/run-status", headers=H).json()["awaiting_allocation"]
        batches = aw["batches"]
        # Split 20/20/20 == 60
        body = {"record_id": rid, "allocations": [
            {"batch_no": batches[0]["batch_no"], "allocate_qty": 20, "consider": True},
            {"batch_no": batches[1]["batch_no"], "allocate_qty": 20, "consider": True},
            {"batch_no": batches[2]["batch_no"], "allocate_qty": 20, "consider": False},
        ]}
        r = requests.post(f"{API}/asn/allocation/confirm", headers=H, json=body)
        assert r.status_code == 200, r.text
        rs = _wait_run_done(H)
        assert rs["processed"] == 1
        rec = _fetch_record(H, TestBatchMultiHappyPath.invoice)
        assert rec["status"] == "Completed", f"got {rec['status']} err={rec.get('error_message')}"
        assert rec["asn_number"].startswith("ASN"), rec["asn_number"]
        TestBatchMultiHappyPath.asn_number = rec["asn_number"]
        events = [e["event"] for e in rec.get("automation_log", [])]
        assert "Batch Details" in events
        assert "Batch Allocation" in events
        assert "Batch Allocated" in events
        # persisted batch_allocations on the record
        allocs = rec.get("batch_allocations") or []
        assert len(allocs) == 3
        total = sum(float(a["allocated_quantity"]) for a in allocs)
        assert abs(total - 60) < 0.001
        # Verify consider flags: 2 Yes, 1 No
        yes_count = sum(1 for a in allocs if a["batch_considerable"] == "Yes")
        assert yes_count == 2, f"expected 2 Yes considerable, got {yes_count}"

    def test_01g_batch_allocations_endpoint(self, H):
        r = requests.get(f"{API}/asn/batch-allocations", headers=H,
                         params={"search": TestBatchMultiHappyPath.asn_number})
        assert r.status_code == 200, r.text
        items = r.json()["items"]
        assert len(items) == 3
        row = items[0]
        for k in ("asn_number", "part_number", "batch_number", "allocated_quantity",
                  "batch_considerable", "created_by", "created_at", "invoice_no"):
            assert k in row, f"missing key {k}: {row}"
        assert row["batch_considerable"] in ("Yes", "No")

    def test_01h_master_dispatch_linked(self, H):
        r = requests.get(f"{API}/master-dispatch/{TestBatchMultiHappyPath.md_id}", headers=H)
        assert r.status_code == 200
        assert r.json().get("asn_number") == TestBatchMultiHappyPath.asn_number

    def test_01i_no_pending_confirm_409(self, H):
        # after run completed, no pending allocation
        r = requests.post(f"{API}/asn/allocation/confirm", headers=H, json={
            "record_id": TestBatchMultiHappyPath.rid, "allocations": [],
        })
        assert r.status_code == 409, r.text

    def test_01j_export_has_batch_column(self, H):
        r = requests.get(f"{API}/asn/export", headers=H)
        assert r.status_code == 200
        assert "spreadsheet" in r.headers.get("content-type", "")
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        assert "Batch Allocations" in headers, f"missing 'Batch Allocations' column: {headers}"

    def test_01k_workflow_shows_batches(self, H):
        r = requests.get(f"{API}/reports/workflow/{TestBatchMultiHappyPath.md_id}", headers=H)
        assert r.status_code == 200, r.text
        steps = r.json().get("steps", [])
        asn_step = next((s for s in steps if s["key"] == "asn"), None)
        assert asn_step, "no asn step in workflow"
        assert asn_step.get("status") == "Completed"
        batches = asn_step.get("batches") or []
        assert len(batches) == 3, f"expected 3 batches on ASN step, got {batches}"
        # each batch row has expected fields
        for b in batches:
            for k in ("part_number", "batch_number", "allocated_quantity", "batch_considerable"):
                assert k in b, f"missing {k}"


# ---------------- 02: BATCH-LOW cancel path ----------------

class TestBatchLowCancel:
    invoice = None
    rid = None

    def test_02a_seed_and_run(self, H):
        _wait_run_idle(H)
        inv = f"TEST/BA-LOW-{int(time.time())}"
        TestBatchLowCancel.invoice = inv
        _seed(H, inv, "BATCH-LOW-XA", 30)
        rec = _import_and_prep(H, inv)
        TestBatchLowCancel.rid = rec["id"]
        r = requests.post(f"{API}/asn/run", headers=H, json={"ids": [rec["id"]]})
        assert r.status_code == 200, r.text

    def test_02b_awaiting_single_short_batch(self, H):
        aw = _wait_awaiting(H)
        assert aw["record_id"] == TestBatchLowCancel.rid
        assert len(aw["batches"]) == 1
        b = aw["batches"][0]
        assert b["available_qty"] < aw["asn_qty"], "available_qty should be < asn_qty for LOW"

    def test_02c_cancel_and_verify_failed_no_retry(self, H):
        r = requests.post(f"{API}/asn/allocation/cancel", headers=H,
                          json={"record_id": TestBatchLowCancel.rid})
        assert r.status_code == 200, r.text
        _wait_run_done(H)
        rec = _fetch_record(H, TestBatchLowCancel.invoice)
        assert rec["status"] == "Failed", f"got {rec['status']}"
        err = (rec.get("error_message") or "").lower()
        assert "cancel" in err, f"error msg: {rec.get('error_message')}"
        assert "after_failure" in (rec.get("screenshots") or {}), "expected after_failure screenshot"
        events = [e["event"] for e in rec.get("automation_log", [])]
        # BatchAllocationError must NOT trigger a Retry event
        assert "Retry" not in events, f"unexpected Retry event on cancel: {events}"

    def test_02d_cancel_no_pending_returns_409(self, H):
        r = requests.post(f"{API}/asn/allocation/cancel", headers=H,
                          json={"record_id": TestBatchLowCancel.rid})
        assert r.status_code == 409, r.text


# ---------------- 03: non-BATCH regression ----------------

class TestNonBatchRegression:
    def test_03_no_pause(self, H):
        _wait_run_idle(H)
        inv = f"TEST/NOBATCH-{int(time.time())}"
        _seed(H, inv, "STD-PART-9001", 10)
        rec = _import_and_prep(H, inv)
        r = requests.post(f"{API}/asn/run", headers=H, json={"ids": [rec["id"]]})
        assert r.status_code == 200, r.text
        # poll — awaiting_allocation should never be set
        start = time.time()
        saw_awaiting = False
        while time.time() - start < 30:
            rs = requests.get(f"{API}/asn/run-status", headers=H).json()
            if rs.get("awaiting_allocation"):
                saw_awaiting = True
                break
            if not rs.get("running"):
                break
            time.sleep(0.5)
        assert not saw_awaiting, "non-BATCH part should not pause for allocation"
        _wait_run_done(H)
        rec = _fetch_record(H, inv)
        assert rec["status"] == "Completed", f"got {rec['status']} err={rec.get('error_message')}"
        assert rec["asn_number"].startswith("ASN")
        assert not (rec.get("batch_allocations") or []), "no batch_allocations expected"
