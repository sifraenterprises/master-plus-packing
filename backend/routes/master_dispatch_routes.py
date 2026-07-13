import io
import math
import re
import uuid
import logging
from io import BytesIO
from datetime import datetime, timezone
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse, FileResponse
from bson import ObjectId
from pymongo import ReturnDocument
from pypdf import PdfReader
from database import db
from models import utcnow
from md_models import MasterDispatch, MasterDispatchInput
from md_ocr import MD_UPLOAD_DIR, launch_batch, next_md_no
from auth import get_current_user, require_admin, log_activity

router = APIRouter(prefix="/master-dispatch", tags=["master-dispatch"])
logger = logging.getLogger(__name__)

MAX_PDF_SIZE = 25 * 1024 * 1024
MAX_FILES = 100
SORT_FIELDS = {"created_at", "invoice_date", "invoice_number", "customer_name", "invoice_total", "dispatch_no", "status"}


def build_md_query(search=None, invoice=None, customer=None, part=None, gstin=None, po=None,
                   eway=None, status=None, verified=None, batch_id=None, date_from=None, date_to=None):
    query = {}
    if search:
        rx = {"$regex": re.escape(search), "$options": "i"}
        query["$or"] = [
            {"dispatch_no": rx}, {"invoice_number": rx}, {"customer_name": rx}, {"customer_code": rx},
            {"gstin": rx}, {"po_number": rx}, {"eway_bill_number": rx}, {"lr_number": rx},
            {"vehicle_number": rx}, {"transporter_name": rx},
            {"items.part_number": rx}, {"items.description": rx},
        ]
    for field, value in [("invoice_number", invoice), ("customer_name", customer), ("gstin", gstin),
                         ("po_number", po), ("eway_bill_number", eway)]:
        if value:
            query[field] = {"$regex": re.escape(value), "$options": "i"}
    if part:
        query["items.part_number"] = {"$regex": re.escape(part), "$options": "i"}
    if status:
        query["status"] = status
    if verified in ("true", "false"):
        query["verified"] = verified == "true"
    if batch_id:
        query["batch_id"] = batch_id
    if date_from or date_to:
        date_q = {}
        if date_from:
            date_q["$gte"] = date_from
        if date_to:
            date_q["$lte"] = date_to
        query["invoice_date"] = date_q
    return query


# ---------- Upload & background OCR ----------

@router.post("/upload")
async def upload_invoices(files: list[UploadFile] = File(...), user: dict = Depends(get_current_user)):
    if not files:
        raise HTTPException(status_code=400, detail="No files provided")
    if len(files) > MAX_FILES:
        raise HTTPException(status_code=400, detail=f"Maximum {MAX_FILES} files per upload")
    payloads = []
    for f in files:
        if not f.filename or not f.filename.lower().endswith(".pdf"):
            raise HTTPException(status_code=400, detail=f"Only PDF files are allowed ({f.filename})")
        content = await f.read()
        if len(content) > MAX_PDF_SIZE:
            raise HTTPException(status_code=400, detail=f"{f.filename} exceeds 25MB limit")
        if not content.startswith(b"%PDF"):
            raise HTTPException(status_code=400, detail=f"{f.filename} is not a valid PDF")
        payloads.append((f.filename, content))

    batch_id = str(uuid.uuid4())
    file_entries = []
    for name, content in payloads:
        file_id = str(uuid.uuid4())
        (MD_UPLOAD_DIR / f"{file_id}.pdf").write_bytes(content)
        try:
            pages = len(PdfReader(BytesIO(content)).pages)
        except Exception:
            pages = 0
        await db.md_uploaded_invoices.insert_one({
            "file_id": file_id, "kind": "original", "original_name": name, "size": len(content),
            "pages": pages, "batch_id": batch_id, "status": "queued", "error": "",
            "uploaded_by": user["username"], "uploaded_at": utcnow().isoformat(),
        })
        file_entries.append({"file_id": file_id, "name": name, "status": "queued",
                             "error": "", "invoices_found": 0, "record_ids": []})
    await db.md_batches.insert_one({
        "batch_id": batch_id, "status": "processing", "total_files": len(file_entries),
        "processed_files": 0, "failed_files": 0, "invoices_created": 0,
        "files": file_entries, "logs": [], "created_by": user["username"],
        "created_at": utcnow().isoformat(), "updated_at": utcnow().isoformat(),
    })
    launch_batch(batch_id)
    await log_activity(user["username"], "md_upload", f"Batch {batch_id}: {len(file_entries)} file(s)", "master_dispatch")
    return {"batch_id": batch_id, "files": len(file_entries), "status": "processing"}


def _serialize_batch(doc: dict) -> dict:
    doc.pop("_id", None)
    return doc


@router.get("/batches")
async def list_batches(page: int = 1, page_size: int = 10, user: dict = Depends(get_current_user)):
    page = max(1, page)
    page_size = min(max(1, page_size), 50)
    total = await db.md_batches.count_documents({})
    docs = await db.md_batches.find({}).sort("created_at", -1).skip((page - 1) * page_size).to_list(page_size)
    return {"items": [_serialize_batch(d) for d in docs], "total": total, "page": page,
            "pages": max(1, math.ceil(total / page_size))}


@router.get("/batches/{batch_id}")
async def get_batch(batch_id: str, user: dict = Depends(get_current_user)):
    doc = await db.md_batches.find_one({"batch_id": batch_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Batch not found")
    return _serialize_batch(doc)


@router.post("/batches/{batch_id}/retry")
async def retry_batch(batch_id: str, user: dict = Depends(get_current_user)):
    doc = await db.md_batches.find_one({"batch_id": batch_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Batch not found")
    failed = [f for f in doc["files"] if f["status"] == "failed"]
    if not failed:
        raise HTTPException(status_code=400, detail="No failed files to retry")
    for f in failed:
        await db.md_batches.update_one(
            {"batch_id": batch_id, "files.file_id": f["file_id"]},
            {"$set": {"files.$.status": "queued", "files.$.error": ""}},
        )
        await db.md_uploaded_invoices.update_one({"file_id": f["file_id"]}, {"$set": {"status": "queued", "error": ""}})
    done = sum(1 for f in doc["files"] if f["status"] == "done")
    await db.md_batches.update_one(
        {"batch_id": batch_id},
        {"$set": {"status": "processing", "processed_files": done, "failed_files": 0,
                  "updated_at": utcnow().isoformat()}},
    )
    launch_batch(batch_id)
    await log_activity(user["username"], "md_retry", f"Batch {batch_id}: retrying {len(failed)} file(s)", "master_dispatch")
    return {"batch_id": batch_id, "retrying": len(failed)}


# ---------- Dashboard stats ----------

@router.get("/stats")
async def md_stats(user: dict = Depends(get_current_user)):
    today = datetime.now(timezone.utc).date().isoformat()
    total = await db.master_dispatch.count_documents({})
    return {
        "total": total,
        "today": await db.master_dispatch.count_documents({"created_at": {"$gte": today}}),
        "pending": await db.master_dispatch.count_documents({"status": "pending"}),
        "ready_for_asn": await db.master_dispatch.count_documents({"status": "ready_for_asn"}),
        "ready_for_eway": await db.master_dispatch.count_documents({"status": "ready_for_eway"}),
        "completed": await db.master_dispatch.count_documents({"status": "completed"}),
        "ocr_errors": await db.md_uploaded_invoices.count_documents({"kind": "original", "status": "failed"}),
        "needs_review": await db.master_dispatch.count_documents({"verified": False}),
    }


# ---------- Stored PDFs ----------

@router.get("/files/{file_id}")
async def get_file(file_id: str, user: dict = Depends(get_current_user)):
    doc = await db.md_uploaded_invoices.find_one({"file_id": file_id})
    path = MD_UPLOAD_DIR / f"{file_id}.pdf"
    if not doc or not path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(str(path), media_type="application/pdf", filename=doc.get("original_name", f"{file_id}.pdf"))


# ---------- Exports ----------

EXPORT_HEADERS = ["Dispatch No", "Invoice No", "Invoice Date", "Customer", "Customer Code", "GSTIN",
                  "PO Number", "PO Date", "Part No", "Description", "HSN", "Qty", "Unit", "Rate", "Amount",
                  "Boxes", "Gross Wt", "Net Wt", "Vehicle", "LR No", "Transporter",
                  "CGST", "SGST", "IGST", "GST Total", "Invoice Total", "E-Way Bill", "IRN", "ACK No", "Status"]


def _export_rows(d: dict):
    base = [d.get("dispatch_no"), d.get("invoice_number"), d.get("invoice_date"), d.get("customer_name"),
            d.get("customer_code"), d.get("gstin"), d.get("po_number"), d.get("po_date")]
    tail = [d.get("boxes"), d.get("gross_weight"), d.get("net_weight"), d.get("vehicle_number"),
            d.get("lr_number"), d.get("transporter_name"), d.get("cgst"), d.get("sgst"), d.get("igst"),
            d.get("gst_total"), d.get("invoice_total"), d.get("eway_bill_number"), d.get("irn"),
            d.get("ack_number"), d.get("status")]
    items = d.get("items") or [{}]
    for it in items:
        yield base + [it.get("part_number", ""), it.get("description", ""), it.get("hsn", ""),
                      it.get("quantity", 0), it.get("unit", ""), it.get("rate", 0), it.get("amount", 0)] + tail


@router.get("/export/excel")
async def md_export_excel(
    search: str = None, invoice: str = None, customer: str = None, part: str = None, gstin: str = None,
    po: str = None, eway: str = None, status: str = None, verified: str = None,
    date_from: str = None, date_to: str = None, user: dict = Depends(get_current_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    query = build_md_query(search, invoice, customer, part, gstin, po, eway, status, verified, None, date_from, date_to)
    docs = await db.master_dispatch.find(query).sort("created_at", -1).to_list(5000)
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Dispatch"
    ws.append(EXPORT_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    for d in docs:
        for row in _export_rows(d):
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    await log_activity(user["username"], "md_export_excel", f"{len(docs)} records", "master_dispatch")
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=master_dispatch.xlsx"},
    )


@router.get("/export/pdf")
async def md_export_pdf(
    search: str = None, invoice: str = None, customer: str = None, part: str = None, gstin: str = None,
    po: str = None, eway: str = None, status: str = None, verified: str = None,
    date_from: str = None, date_to: str = None, user: dict = Depends(get_current_user),
):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    query = build_md_query(search, invoice, customer, part, gstin, po, eway, status, verified, None, date_from, date_to)
    docs = await db.master_dispatch.find(query).sort("created_at", -1).to_list(2000)
    cols = [("Dispatch No", "dispatch_no"), ("Invoice No", "invoice_number"), ("Date", "invoice_date"),
            ("Customer", "customer_name"), ("GSTIN", "gstin"), ("Items", None), ("Invoice Total", "invoice_total"),
            ("E-Way Bill", "eway_bill_number"), ("Status", "status")]
    buf = io.BytesIO()
    pdf_doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=15 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Grewal Engineering Work — Master Dispatch Report", styles["Title"]),
        Paragraph(f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} | Records: {len(docs)}", styles["Normal"]),
        Spacer(1, 8),
    ]
    data = [[c[0] for c in cols]]
    for d in docs:
        data.append([str(len(d.get("items") or [])) if c[1] is None else str(d.get(c[1], ""))[:32] for c in cols])
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F97316")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
    ]))
    elements.append(table)
    pdf_doc.build(elements)
    buf.seek(0)
    await log_activity(user["username"], "md_export_pdf", f"{len(docs)} records", "master_dispatch")
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=master_dispatch_report.pdf"},
    )


# ---------- CRUD ----------

@router.post("")
async def create_md(body: MasterDispatchInput, user: dict = Depends(get_current_user)):
    record = MasterDispatch(**body.model_dump(), dispatch_no=await next_md_no(),
                            ocr_status="manual", created_by=user["username"])
    result = await db.master_dispatch.insert_one(record.to_mongo())
    await log_activity(user["username"], "md_created", record.dispatch_no, "master_dispatch")
    doc = await db.master_dispatch.find_one({"_id": result.inserted_id})
    return MasterDispatch.from_mongo(doc).model_dump()


@router.get("")
async def list_md(
    search: str = None, invoice: str = None, customer: str = None, part: str = None, gstin: str = None,
    po: str = None, eway: str = None, status: str = None, verified: str = None, batch_id: str = None,
    date_from: str = None, date_to: str = None,
    sort_by: str = "created_at", sort_dir: str = "desc",
    page: int = 1, page_size: int = 25, user: dict = Depends(get_current_user),
):
    query = build_md_query(search, invoice, customer, part, gstin, po, eway, status, verified, batch_id, date_from, date_to)
    sort_by = sort_by if sort_by in SORT_FIELDS else "created_at"
    direction = 1 if sort_dir == "asc" else -1
    page = max(1, page)
    page_size = min(max(1, page_size), 100)
    total = await db.master_dispatch.count_documents(query)
    docs = (
        await db.master_dispatch.find(query)
        .sort(sort_by, direction)
        .skip((page - 1) * page_size)
        .to_list(page_size)
    )
    return {
        "items": [MasterDispatch.from_mongo(d).model_dump() for d in docs],
        "total": total, "page": page, "page_size": page_size,
        "pages": max(1, math.ceil(total / page_size)),
    }


@router.get("/{record_id}")
async def get_md(record_id: str, user: dict = Depends(get_current_user)):
    if not ObjectId.is_valid(record_id):
        raise HTTPException(status_code=400, detail="Invalid record ID")
    doc = await db.master_dispatch.find_one({"_id": ObjectId(record_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Record not found")
    return MasterDispatch.from_mongo(doc).model_dump()


@router.put("/{record_id}")
async def update_md(record_id: str, body: MasterDispatchInput, user: dict = Depends(get_current_user)):
    if not ObjectId.is_valid(record_id):
        raise HTTPException(status_code=400, detail="Invalid record ID")
    update = body.model_dump()
    update["items"] = [i for i in update["items"]]
    update["updated_at"] = utcnow().isoformat()
    result = await db.master_dispatch.find_one_and_update(
        {"_id": ObjectId(record_id)}, {"$set": update}, return_document=ReturnDocument.AFTER
    )
    if not result:
        raise HTTPException(status_code=404, detail="Record not found")
    await log_activity(user["username"], "md_updated", result.get("dispatch_no", record_id), "master_dispatch")
    return MasterDispatch.from_mongo(result).model_dump()


@router.delete("/{record_id}")
async def delete_md(record_id: str, user: dict = Depends(require_admin)):
    if not ObjectId.is_valid(record_id):
        raise HTTPException(status_code=400, detail="Invalid record ID")
    doc = await db.master_dispatch.find_one_and_delete({"_id": ObjectId(record_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Record not found")
    await log_activity(user["username"], "md_deleted", doc.get("dispatch_no", record_id), "master_dispatch")
    return {"message": "Record deleted"}


@router.post("/{record_id}/duplicate")
async def duplicate_md(record_id: str, user: dict = Depends(get_current_user)):
    if not ObjectId.is_valid(record_id):
        raise HTTPException(status_code=400, detail="Invalid record ID")
    doc = await db.master_dispatch.find_one({"_id": ObjectId(record_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Record not found")
    doc.pop("_id", None)
    doc["dispatch_no"] = await next_md_no()
    doc["status"] = "pending"
    doc["created_by"] = user["username"]
    doc["created_at"] = utcnow().isoformat()
    doc["updated_at"] = utcnow().isoformat()
    result = await db.master_dispatch.insert_one(doc)
    await log_activity(user["username"], "md_duplicated", f"→ {doc['dispatch_no']}", "master_dispatch")
    new_doc = await db.master_dispatch.find_one({"_id": result.inserted_id})
    return MasterDispatch.from_mongo(new_doc).model_dump()
