import io
import csv
import re
from datetime import datetime, timezone, timedelta
from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from environment import env_list_filter
from database import db
from models import utcnow
from auth import get_current_user, log_activity

router = APIRouter(prefix="/reports", tags=["reports"])

STATUS_KEYS = ("packing_status", "asn_status", "eway_status", "vendor_ack_status", "pdi_status")
STATUS_VALUES = ("Completed", "Pending", "Failed")
SORT_FIELDS = {"invoice_number", "invoice_date", "customer_name", "plant", "quantity", "dispatch_date",
               "transporter_name", "created_at", *STATUS_KEYS}

ERP_COLUMNS = [
    ("invoice_number", "Invoice Number"), ("invoice_date", "Invoice Date"), ("customer_name", "Customer"),
    ("plant", "Plant"), ("part_numbers", "Part Number"), ("quantity", "Quantity"),
    ("packing_status", "Packing Slip"), ("asn_status", "ASN"), ("eway_status", "E-Way Bill"),
    ("vendor_ack_status", "Vendor Ack"), ("pdi_status", "PDI"), ("dispatch_date", "Dispatch Date"),
    ("transporter_name", "Transporter"), ("vehicle_number", "Vehicle"), ("po_number", "PO Number"),
    ("asn_no", "ASN Number"), ("eway_bill_number", "E-Way Bill No"), ("packing_slip_no", "Packing Slip No"),
    ("boxes", "Boxes"),
]


@router.get("/summary")
async def summary(user: dict = Depends(get_current_user)):
    total = await db.dispatch_entries.count_documents({})
    month_prefix = datetime.now(timezone.utc).strftime("%Y-%m")
    this_month = await db.dispatch_entries.count_documents({"created_at": {"$regex": f"^{month_prefix}"}})
    customers = await db.dispatch_entries.distinct("customer_name")
    pipeline = [{"$group": {"_id": None, "total": {"$sum": "$total_value"}}}]
    agg = await db.dispatch_entries.aggregate(pipeline).to_list(1)
    total_value = agg[0]["total"] if agg else 0
    pdfs = await db.uploaded_pdfs.count_documents({})
    return {
        "total_dispatches": total,
        "this_month": this_month,
        "unique_customers": len([c for c in customers if c]),
        "total_value": round(total_value, 2),
        "pdfs_uploaded": pdfs,
    }


# ---------- ERP pipeline (master_dispatch + module status joins) ----------

def _rx(v):
    return {"$regex": re.escape(v), "$options": "i"}


def build_erp_pipeline(p: dict, env_filter: dict = None):
    stages = [{"$addFields": {"dispatch_date": {"$substrCP": [{"$ifNull": ["$created_at", ""]}, 0, 10]}}}]
    pre = dict(env_filter or {})
    for param, field in (("invoice", "invoice_number"), ("customer", "customer_name"), ("vendor", "customer_code"),
                         ("plant", "plant"), ("transporter", "transporter_name"), ("vehicle", "vehicle_number"),
                         ("eway", "eway_bill_number"), ("po", "po_number"), ("part", "items.part_number"),
                         ("description", "items.description")):
        if p.get(param):
            pre[field] = _rx(p[param])
    if p.get("search"):
        rx = _rx(p["search"])
        pre["$or"] = [{f: rx} for f in ("invoice_number", "customer_name", "plant", "transporter_name",
                                        "po_number", "eway_bill_number", "asn_number", "vehicle_number",
                                        "items.part_number", "items.description", "dispatch_no")]
    for param, field, op in (("inv_from", "invoice_date", "$gte"), ("inv_to", "invoice_date", "$lte"),
                             ("dispatch_from", "dispatch_date", "$gte"), ("dispatch_to", "dispatch_date", "$lte")):
        if p.get(param):
            pre.setdefault(field, {})[op] = p[param]
    if pre:
        stages.append({"$match": pre})
    stages += [
        {"$lookup": {"from": "packing_slips", "localField": "invoice_number",
                     "foreignField": "invoice_number", "as": "_pk"}},
        {"$lookup": {"from": "asn_creation", "let": {"mid": {"$toString": "$_id"}},
                     "pipeline": [{"$match": {"$expr": {"$eq": ["$master_dispatch_id", "$$mid"]}}},
                                  {"$project": {"status": 1, "asn_number": 1, "completed_at": 1,
                                                "created_at": 1, "error_message": 1}}], "as": "_asn"}},
        {"$lookup": {"from": "eway_submissions", "let": {"mid": {"$toString": "$_id"}},
                     "pipeline": [{"$match": {"$expr": {"$eq": ["$record_id", "$$mid"]}}},
                                  {"$project": {"status": 1, "completed_time": 1, "error": 1, "updated_at": 1}}],
                     "as": "_ew"}},
        {"$lookup": {"from": "vendor_eway_acknowledgement", "let": {"mid": {"$toString": "$_id"}},
                     "pipeline": [{"$match": {"$expr": {"$eq": ["$dispatch_id", "$$mid"]}}},
                                  {"$project": {"status": 1, "ack_date": 1, "ack_time": 1,
                                                "portal_message": 1, "updated_at": 1}}], "as": "_va"}},
        {"$lookup": {"from": "pdi_reports", "let": {"mid": {"$toString": "$_id"}},
                     "pipeline": [{"$match": {"$expr": {"$eq": ["$master_dispatch_id", "$$mid"]}}},
                                  {"$project": {"report_no": 1, "created_at": 1}}], "as": "_pdi"}},
        {"$addFields": {
            "packing_status": {"$cond": [{"$gt": [{"$size": "$_pk"}, 0]}, "Completed", "Pending"]},
            "packing_slip_no": {"$ifNull": [{"$first": "$_pk.lot_number"}, ""]},
            "asn_status": {"$switch": {"branches": [
                {"case": {"$ne": [{"$ifNull": ["$asn_number", ""]}, ""]}, "then": "Completed"},
                {"case": {"$eq": [{"$first": "$_asn.status"}, "Completed"]}, "then": "Completed"},
                {"case": {"$eq": [{"$first": "$_asn.status"}, "Failed"]}, "then": "Failed"},
            ], "default": "Pending"}},
            "asn_no": {"$cond": [{"$ne": [{"$ifNull": ["$asn_number", ""]}, ""]}, "$asn_number",
                                 {"$ifNull": [{"$first": "$_asn.asn_number"}, ""]}]},
            "eway_status": {"$switch": {"branches": [
                {"case": {"$eq": [{"$first": "$_ew.status"}, "Completed"]}, "then": "Completed"},
                {"case": {"$eq": [{"$first": "$_ew.status"}, "Failed"]}, "then": "Failed"},
            ], "default": "Pending"}},
            "vendor_ack_status": {"$switch": {"branches": [
                {"case": {"$eq": [{"$first": "$_va.status"}, "Completed"]}, "then": "Completed"},
                {"case": {"$eq": [{"$first": "$_va.status"}, "Failed"]}, "then": "Failed"},
                {"case": {"$eq": [{"$ifNull": ["$vendor_ack_status", ""]}, "Completed"]}, "then": "Completed"},
            ], "default": "Pending"}},
            "pdi_status": {"$cond": [{"$gt": [{"$size": "$_pdi"}, 0]}, "Completed", "Pending"]},
            "pdi_report_no": {"$ifNull": [{"$first": "$_pdi.report_no"}, ""]},
            "quantity": {"$sum": "$items.quantity"},
            "part_numbers": {"$reduce": {"input": {"$ifNull": ["$items.part_number", []]}, "initialValue": "",
                                         "in": {"$cond": [{"$eq": ["$$value", ""]}, "$$this",
                                                          {"$concat": ["$$value", ", ", "$$this"]}]}}},
        }},
    ]
    post = {}
    for key in STATUS_KEYS:
        if p.get(key) in STATUS_VALUES:
            post[key] = p[key]
    if p.get("asn"):
        post["asn_no"] = _rx(p["asn"])
    if p.get("packing_slip"):
        post["packing_slip_no"] = _rx(p["packing_slip"])
    if post:
        stages.append({"$match": post})
    stages.append({"$project": {"_pk": 0, "_asn": 0, "_ew": 0, "_va": 0, "_pdi": 0, "confidence": 0,
                                "automation_log": 0, "low_confidence_fields": 0, "items": 0}})
    return stages


def erp_params(search=None, invoice=None, customer=None, vendor=None, plant=None, transporter=None,
               vehicle=None, packing_slip=None, asn=None, eway=None, po=None, part=None, description=None,
               inv_from=None, inv_to=None, dispatch_from=None, dispatch_to=None,
               packing_status=None, asn_status=None, eway_status=None, vendor_ack_status=None, pdi_status=None):
    return {k: v for k, v in locals().items() if v}


def _row(d: dict) -> dict:
    d = dict(d)
    d["id"] = str(d.pop("_id"))
    return d


@router.get("/erp")
async def erp_report(sort_by: str = "created_at", sort_dir: str = "desc", page: int = 1, page_size: int = 25,
                     user: dict = Depends(get_current_user), p: dict = Depends(erp_params)):
    pipeline = build_erp_pipeline(p, await env_list_filter())
    sort_by = sort_by if sort_by in SORT_FIELDS else "created_at"
    page, page_size = max(1, page), min(max(1, page_size), 200)
    pipeline += [
        {"$sort": {sort_by: 1 if sort_dir == "asc" else -1, "_id": 1}},
        {"$facet": {"rows": [{"$skip": (page - 1) * page_size}, {"$limit": page_size}],
                    "total": [{"$count": "n"}]}},
    ]
    res = (await db.master_dispatch.aggregate(pipeline).to_list(1))[0]
    total = res["total"][0]["n"] if res["total"] else 0
    return {"items": [_row(d) for d in res["rows"]], "total": total, "page": page,
            "page_size": page_size, "pages": max(1, -(-total // page_size))}


@router.get("/kpis")
async def erp_kpis(user: dict = Depends(get_current_user)):
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    month = today[:7]
    group = {"_id": None, "n": {"$sum": 1}}
    for key in STATUS_KEYS:
        short = key.replace("_status", "")
        group[f"pending_{short}"] = {"$sum": {"$cond": [{"$ne": [f"${key}", "Completed"]}, 1, 0]}}
        group[f"completed_{short}"] = {"$sum": {"$cond": [{"$eq": [f"${key}", "Completed"]}, 1, 0]}}
    status_agg = await db.master_dispatch.aggregate(build_erp_pipeline({}, await env_list_filter()) + [{"$group": group}]).to_list(1)
    counts = status_agg[0] if status_agg else {}
    counts.pop("_id", None)

    async def day_stats(prefix):
        agg = await db.master_dispatch.aggregate([
            {"$match": {"created_at": {"$regex": f"^{re.escape(prefix)}"}}},
            {"$group": {"_id": None, "n": {"$sum": 1}, "boxes": {"$sum": {"$ifNull": ["$boxes", 0]}}}},
        ]).to_list(1)
        return (agg[0]["n"], agg[0]["boxes"]) if agg else (0, 0)

    today_n, today_boxes = await day_stats(today)
    month_n, month_boxes = await day_stats(month)
    return {"today_dispatches": today_n, "today_boxes": today_boxes,
            "month_dispatches": month_n, "month_boxes": month_boxes,
            "total_dispatches": counts.get("n", 0), **counts, "today": today}


@router.get("/charts")
async def erp_charts(user: dict = Depends(get_current_user)):
    def top_group(field):
        return [{"$match": {field: {"$nin": ["", None]}}},
                {"$group": {"_id": f"${field}", "count": {"$sum": 1}, "boxes": {"$sum": {"$ifNull": ["$boxes", 0]}}}},
                {"$sort": {"count": -1}}, {"$limit": 8},
                {"$project": {"name": "$_id", "count": 1, "boxes": 1, "_id": 0}}]

    by_month = await db.master_dispatch.aggregate([
        {"$match": {"invoice_date": {"$regex": r"^\d{4}-\d{2}"}}},
        {"$group": {"_id": {"$substrCP": ["$invoice_date", 0, 7]}, "count": {"$sum": 1},
                    "boxes": {"$sum": {"$ifNull": ["$boxes", 0]}}}},
        {"$sort": {"_id": 1}}, {"$limit": 24},
        {"$project": {"name": "$_id", "count": 1, "boxes": 1, "_id": 0}},
    ]).to_list(24)
    since = (datetime.now(timezone.utc) - timedelta(days=30)).strftime("%Y-%m-%d")
    boxes_per_day = await db.master_dispatch.aggregate([
        {"$match": {"invoice_date": {"$gte": since}}},
        {"$group": {"_id": "$invoice_date", "boxes": {"$sum": {"$ifNull": ["$boxes", 0]}}, "count": {"$sum": 1}}},
        {"$sort": {"_id": 1}},
        {"$project": {"name": "$_id", "boxes": 1, "count": 1, "_id": 0}},
    ]).to_list(60)
    by_customer = await db.master_dispatch.aggregate(top_group("customer_name")).to_list(8)
    by_plant = await db.master_dispatch.aggregate(top_group("plant")).to_list(8)
    by_transporter = await db.master_dispatch.aggregate(top_group("transporter_name")).to_list(8)
    kpis = await erp_kpis(user)
    total = max(1, kpis.get("n", 0))
    completion = {short: round(kpis.get(f"completed_{short}", 0) * 100 / total, 1)
                  for short in ("asn", "eway", "vendor_ack")}
    return {"by_month": by_month[-12:], "by_customer": by_customer, "by_plant": by_plant,
            "by_transporter": by_transporter, "boxes_per_day": boxes_per_day, "completion": completion,
            "total_records": kpis.get("n", 0)}


@router.get("/group")
async def erp_group(by: str = "customer", user: dict = Depends(get_current_user)):
    field = {"customer": "customer_name", "plant": "plant", "transporter": "transporter_name",
             "month": None}.get(by)
    if by == "month":
        group_id = {"$substrCP": [{"$ifNull": ["$invoice_date", ""]}, 0, 7]}
    elif field:
        group_id = {"$ifNull": [f"${field}", ""]}
    else:
        raise HTTPException(status_code=400, detail="by must be customer|plant|transporter|month")
    rows = await db.master_dispatch.aggregate([
        {"$group": {"_id": group_id, "dispatches": {"$sum": 1},
                    "boxes": {"$sum": {"$ifNull": ["$boxes", 0]}},
                    "value": {"$sum": {"$ifNull": ["$invoice_total", 0]}}}},
        {"$sort": {"dispatches": -1} if by != "month" else {"_id": 1}},
        {"$project": {"name": {"$cond": [{"$eq": ["$_id", ""]}, "(blank)", "$_id"]},
                      "dispatches": 1, "boxes": 1, "value": {"$round": ["$value", 2]}, "_id": 0}},
    ]).to_list(500)
    return {"by": by, "rows": rows,
            "totals": {"dispatches": sum(r["dispatches"] for r in rows),
                       "boxes": sum(r["boxes"] for r in rows),
                       "value": round(sum(r["value"] for r in rows), 2)}}


# ---------- Exports ----------

@router.get("/erp/export")
async def erp_export(format: str = "excel", columns: str = "", sort_by: str = "created_at", sort_dir: str = "desc",
                     user: dict = Depends(get_current_user), p: dict = Depends(erp_params)):
    keys = [k for k in columns.split(",") if k in dict(ERP_COLUMNS)] or [k for k, _ in ERP_COLUMNS[:12]]
    labels = [dict(ERP_COLUMNS)[k] for k in keys]
    sort_by = sort_by if sort_by in SORT_FIELDS else "created_at"
    pipeline = build_erp_pipeline(p, await env_list_filter()) + [{"$sort": {sort_by: 1 if sort_dir == "asc" else -1, "_id": 1}}, {"$limit": 10000}]
    docs = await db.master_dispatch.aggregate(pipeline).to_list(10000)
    await log_activity(user["username"], f"erp_report_export_{format}", f"{len(docs)} records", "reports")

    if format == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(labels)
        for d in docs:
            writer.writerow([d.get(k, "") for k in keys])
        return StreamingResponse(io.BytesIO(buf.getvalue().encode("utf-8-sig")), media_type="text/csv",
                                 headers={"Content-Disposition": "attachment; filename=erp_report.csv"})

    if format == "pdf":
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=12 * mm, bottomMargin=12 * mm)
        styles = getSampleStyleSheet()
        data = [labels] + [[str(d.get(k, ""))[:28] for k in keys] for d in docs[:2000]]
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F97316")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 6.5),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
        ]))
        doc.build([Paragraph("Grewal Engineering Works — ERP Dispatch Report", styles["Title"]),
                   Paragraph(f"Generated: {datetime.now(timezone.utc):%Y-%m-%d %H:%M UTC} | Records: {len(docs)}",
                             styles["Normal"]), Spacer(1, 6), table])
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf",
                                 headers={"Content-Disposition": "attachment; filename=erp_report.pdf"})

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "ERP Report"
    ws.append(labels)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    for d in docs:
        ws.append([d.get(k, "") for k in keys])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=erp_report.xlsx"})


# ---------- Drill-down workflow ----------

@router.get("/workflow/{md_id}")
async def dispatch_workflow(md_id: str, user: dict = Depends(get_current_user)):
    if not ObjectId.is_valid(md_id):
        raise HTTPException(status_code=400, detail="Invalid ID")
    md = await db.master_dispatch.find_one({"_id": ObjectId(md_id)})
    if not md:
        raise HTTPException(status_code=404, detail="Dispatch not found")
    pk = await db.packing_slips.find_one({"invoice_number": md.get("invoice_number", "")}, sort=[("created_at", -1)])
    asn = await db.asn_creation.find_one({"master_dispatch_id": md_id})
    ew = await db.eway_submissions.find_one({"record_id": md_id})
    va = await db.vendor_eway_acknowledgement.find_one({"dispatch_id": md_id})
    pdi = await db.pdi_reports.find_one({"master_dispatch_id": md_id}, sort=[("created_at", -1)])
    batch_allocs = await db.asn_batch_allocations.find({"dispatch_id": md_id}).sort("created_at", -1).to_list(50)
    file_id = md.get("split_file_id") or md.get("source_file_id") or ""
    steps = [
        {"key": "master_dispatch", "label": "Master Dispatch", "status": "Completed",
         "doc_no": md.get("dispatch_no", ""), "timestamp": md.get("created_at", ""),
         "detail": f"Invoice {md.get('invoice_number', '')} · {md.get('customer_name', '')}",
         "download": f"/master-dispatch/files/{file_id}" if file_id else ""},
        {"key": "packing", "label": "Packing Slip", "status": "Completed" if pk else "Pending",
         "doc_no": (pk or {}).get("lot_number", ""), "timestamp": (pk or {}).get("created_at", ""),
         "detail": f"{(pk or {}).get('boxes', '')} boxes" if pk else "No packing slip saved for this invoice",
         "download": ""},
        {"key": "asn", "label": "ASN Creation",
         "status": "Completed" if (md.get("asn_number") or (asn or {}).get("status") == "Completed")
         else ("Failed" if (asn or {}).get("status") == "Failed" else "Pending"),
         "doc_no": md.get("asn_number") or (asn or {}).get("asn_number", ""),
         "timestamp": (asn or {}).get("completed_at") or (asn or {}).get("updated_at", ""),
         "detail": (asn or {}).get("error_message") or ("ASN captured from portal" if md.get("asn_number") else "Awaiting ASN creation"),
         "batches": [{"part_number": b.get("part_number", ""), "batch_number": b.get("batch_number", ""),
                      "allocated_quantity": b.get("allocated_quantity", 0),
                      "batch_considerable": b.get("batch_considerable", "")} for b in batch_allocs],
         "download": ""},
        {"key": "eway", "label": "E-Way Bill",
         "status": (ew or {}).get("status") if (ew or {}).get("status") in STATUS_VALUES else "Pending",
         "doc_no": md.get("eway_bill_number", ""), "timestamp": (ew or {}).get("completed_time") or (ew or {}).get("updated_at", ""),
         "detail": (ew or {}).get("error") or ("Submitted to TAFE portal" if (ew or {}).get("status") == "Completed" else "Awaiting portal submission"),
         "download": ""},
        {"key": "vendor_ack", "label": "Vendor Acknowledgement",
         "status": (va or {}).get("status") if (va or {}).get("status") in STATUS_VALUES else
         ("Completed" if md.get("vendor_ack_status") == "Completed" else "Pending"),
         "doc_no": (va or {}).get("asn_number", ""),
         "timestamp": f"{(va or {}).get('ack_date', '')} {(va or {}).get('ack_time', '')}".strip(),
         "detail": (va or {}).get("portal_message", "") or "Awaiting acknowledgement",
         "download": ""},
        {"key": "pdi", "label": "PDI Report", "status": "Completed" if pdi else "Pending",
         "doc_no": (pdi or {}).get("report_no", ""), "timestamp": (pdi or {}).get("created_at", ""),
         "detail": f"{(pdi or {}).get('part_name', '')} · Item {(pdi or {}).get('item_code', '')}" if pdi
         else "No PDI report generated yet",
         "download": f"/pdi/reports/{str(pdi['_id'])}/pdf" if pdi else ""},
    ]
    return {"dispatch": {"id": md_id, "dispatch_no": md.get("dispatch_no", ""),
                         "invoice_number": md.get("invoice_number", ""), "invoice_date": md.get("invoice_date", ""),
                         "customer_name": md.get("customer_name", ""), "plant": md.get("plant", ""),
                         "transporter_name": md.get("transporter_name", ""), "vehicle_number": md.get("vehicle_number", ""),
                         "po_number": md.get("po_number", ""), "boxes": md.get("boxes", 0),
                         "invoice_total": md.get("invoice_total", 0)},
            "steps": steps}


# ---------- Saved report views ----------

class ViewInput(BaseModel):
    name: str = Field(min_length=1, max_length=80)
    filters: dict = {}
    columns: list[str] = []
    scope: str = Field(default="personal", pattern="^(personal|shared)$")


@router.get("/views")
async def list_views(user: dict = Depends(get_current_user)):
    docs = await db.report_views.find({"$or": [{"owner": user["username"]}, {"scope": "shared"}]}) \
        .sort("name", 1).to_list(200)
    pref = await db.report_view_prefs.find_one({"username": user["username"]})
    return {"views": [_row(d) for d in docs], "default_view_id": (pref or {}).get("view_id", "")}


@router.post("/views")
async def save_view(body: ViewInput, user: dict = Depends(get_current_user)):
    if body.scope == "shared" and user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admins can create shared report templates")
    filters = {k: v for k, v in body.filters.items() if isinstance(v, str) and v}
    doc = {"name": body.name.strip(), "filters": filters, "columns": body.columns[:30],
           "scope": body.scope, "owner": user["username"], "created_at": utcnow().isoformat()}
    result = await db.report_views.insert_one(doc)
    await log_activity(user["username"], "report_view_saved", f"{body.name} ({body.scope})", "reports")
    return _row(await db.report_views.find_one({"_id": result.inserted_id}))


@router.delete("/views/{view_id}")
async def delete_view(view_id: str, user: dict = Depends(get_current_user)):
    if not ObjectId.is_valid(view_id):
        raise HTTPException(status_code=400, detail="Invalid ID")
    doc = await db.report_views.find_one({"_id": ObjectId(view_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="View not found")
    if doc["owner"] != user["username"] and user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="You can only delete your own report views")
    await db.report_views.delete_one({"_id": doc["_id"]})
    await db.report_view_prefs.delete_many({"view_id": view_id})
    await log_activity(user["username"], "report_view_deleted", doc["name"], "reports")
    return {"message": "View deleted"}


@router.post("/views/{view_id}/default")
async def set_default_view(view_id: str, user: dict = Depends(get_current_user)):
    if view_id != "none":
        if not ObjectId.is_valid(view_id) or not await db.report_views.find_one({"_id": ObjectId(view_id)}):
            raise HTTPException(status_code=404, detail="View not found")
    await db.report_view_prefs.update_one(
        {"username": user["username"]},
        {"$set": {"view_id": "" if view_id == "none" else view_id, "updated_at": utcnow().isoformat()}},
        upsert=True)
    return {"default_view_id": "" if view_id == "none" else view_id}
