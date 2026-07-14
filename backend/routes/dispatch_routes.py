import os
import io
import json
import math
import re
import uuid
import logging
from pathlib import Path
from datetime import datetime, timezone
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from bson import ObjectId
from pymongo import ReturnDocument
from database import db
from models import DispatchEntry, DispatchEntryInput, BulkDispatchInput, utcnow
from auth import get_current_user, log_activity

router = APIRouter(prefix="/dispatch", tags=["dispatch"])
logger = logging.getLogger(__name__)

UPLOAD_DIR = Path(__file__).parent.parent / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)
MAX_PDF_SIZE = 10 * 1024 * 1024

EXTRACTION_PROMPT = """Extract all invoice/dispatch data from this PDF invoice. Return ONLY valid JSON (no markdown, no explanation) in exactly this structure:
{
  "invoice_number": "", "invoice_date": "YYYY-MM-DD", "customer_name": "", "customer_code": "",
  "po_number": "", "vendor_name": "", "vehicle": "", "dispatch_date": "YYYY-MM-DD", "gst": "",
  "items": [{"part_number": "", "part_description": "", "quantity": 0, "unit": "", "rate": 0, "total_value": 0}]
}
Rules: quantities/rates/totals must be numbers. Dates in YYYY-MM-DD. Use empty string for missing text fields and 0 for missing numbers. "gst" is the GST amount or percentage as string. Include every line item in "items"."""


async def next_dispatch_id() -> str:
    counter = await db.counters.find_one_and_update(
        {"_id": "dispatch"}, {"$inc": {"seq": 1}}, upsert=True, return_document=ReturnDocument.AFTER
    )
    return f"GEW-DSP-{counter['seq']:05d}"


def build_query(search=None, invoice=None, part=None, customer=None, date_from=None, date_to=None):
    query = {}
    if search:
        rx = {"$regex": re.escape(search), "$options": "i"}
        query["$or"] = [
            {"invoice_number": rx}, {"customer_name": rx}, {"part_number": rx},
            {"part_description": rx}, {"dispatch_id": rx}, {"po_number": rx}, {"vendor_name": rx},
        ]
    if invoice:
        query["invoice_number"] = {"$regex": re.escape(invoice), "$options": "i"}
    if part:
        query["part_number"] = {"$regex": re.escape(part), "$options": "i"}
    if customer:
        query["customer_name"] = {"$regex": re.escape(customer), "$options": "i"}
    if date_from or date_to:
        date_q = {}
        if date_from:
            date_q["$gte"] = date_from
        if date_to:
            date_q["$lte"] = date_to
        query["invoice_date"] = date_q
    return query


@router.post("/extract")
async def extract_invoice(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Only PDF files are allowed")
    content = await file.read()
    if len(content) > MAX_PDF_SIZE:
        raise HTTPException(status_code=400, detail="File too large (max 10MB)")
    if not content.startswith(b"%PDF"):
        raise HTTPException(status_code=400, detail="Invalid PDF file")

    file_id = str(uuid.uuid4())
    path = UPLOAD_DIR / f"{file_id}.pdf"
    path.write_bytes(content)
    await db.uploaded_pdfs.insert_one({
        "file_id": file_id, "original_name": file.filename, "size": len(content),
        "uploaded_by": user["username"], "uploaded_at": utcnow().isoformat(),
    })

    try:
        from google import genai
        from google.genai import types
        client = genai.Client(api_key=os.environ["GEMINI_API_KEY"])
        resp = await client.aio.models.generate_content(
            model=os.environ.get("GEMINI_MODEL", "gemini-flash-latest"),
            contents=[types.Part.from_bytes(data=content, mime_type="application/pdf"),
                      "You are an expert invoice data extraction engine for an Indian engineering company. You always return strict JSON.\n\n" + EXTRACTION_PROMPT],
        )
        raw = resp.text or ""
    except Exception as e:
        logger.error(f"AI extraction failed: {e}")
        await log_activity(user["username"], "extraction_failed", str(e)[:200], "dispatch")
        raise HTTPException(status_code=502, detail="AI extraction failed. You can still create the entry manually.")

    match = re.search(r"\{.*\}", raw, re.DOTALL)
    if not match:
        raise HTTPException(status_code=502, detail="Could not parse extracted data. Please enter details manually.")
    try:
        data = json.loads(match.group(0))
    except json.JSONDecodeError:
        raise HTTPException(status_code=502, detail="Could not parse extracted data. Please enter details manually.")

    header = {k: str(data.get(k, "") or "") for k in [
        "invoice_number", "invoice_date", "customer_name", "customer_code",
        "po_number", "vendor_name", "vehicle", "dispatch_date", "gst"]}
    items = data.get("items") or [{}]
    entries = []
    for item in items:
        entries.append({
            **header,
            "part_number": str(item.get("part_number", "") or ""),
            "part_description": str(item.get("part_description", "") or ""),
            "quantity": float(item.get("quantity") or 0),
            "unit": str(item.get("unit", "") or ""),
            "rate": float(item.get("rate") or 0),
            "total_value": float(item.get("total_value") or 0),
            "remarks": "",
            "pdf_id": file_id,
        })
    await log_activity(user["username"], "invoice_extracted", f"{file.filename} -> {len(entries)} items", "dispatch")
    return {"pdf_id": file_id, "filename": file.filename, "entries": entries}


@router.post("")
async def create_entry(body: DispatchEntryInput, user: dict = Depends(get_current_user)):
    entry = DispatchEntry(**body.model_dump(), dispatch_id=await next_dispatch_id(), created_by=user["username"])
    result = await db.dispatch_entries.insert_one(entry.to_mongo())
    await log_activity(user["username"], "dispatch_created", entry.dispatch_id, "dispatch")
    doc = await db.dispatch_entries.find_one({"_id": result.inserted_id})
    return DispatchEntry.from_mongo(doc).model_dump()


@router.post("/bulk")
async def create_bulk(body: BulkDispatchInput, user: dict = Depends(get_current_user)):
    if not body.entries:
        raise HTTPException(status_code=400, detail="No entries provided")
    created = []
    for item in body.entries:
        entry = DispatchEntry(**item.model_dump(), dispatch_id=await next_dispatch_id(), created_by=user["username"])
        await db.dispatch_entries.insert_one(entry.to_mongo())
        created.append(entry.dispatch_id)
    await log_activity(user["username"], "dispatch_bulk_created", ", ".join(created), "dispatch")
    return {"created": created, "count": len(created)}


@router.get("")
async def list_entries(
    search: str = None, invoice: str = None, part: str = None, customer: str = None,
    date_from: str = None, date_to: str = None, page: int = 1, page_size: int = 25,
    user: dict = Depends(get_current_user),
):
    query = build_query(search, invoice, part, customer, date_from, date_to)
    page = max(1, page)
    page_size = min(max(1, page_size), 100)
    total = await db.dispatch_entries.count_documents(query)
    docs = (
        await db.dispatch_entries.find(query)
        .sort("created_at", -1)
        .skip((page - 1) * page_size)
        .to_list(page_size)
    )
    return {
        "items": [DispatchEntry.from_mongo(d).model_dump() for d in docs],
        "total": total,
        "page": page,
        "page_size": page_size,
        "pages": max(1, math.ceil(total / page_size)),
    }


@router.put("/{entry_id}")
async def update_entry(entry_id: str, body: DispatchEntryInput, user: dict = Depends(get_current_user)):
    if not ObjectId.is_valid(entry_id):
        raise HTTPException(status_code=400, detail="Invalid entry ID")
    update = body.model_dump()
    update["updated_at"] = utcnow().isoformat()
    result = await db.dispatch_entries.find_one_and_update(
        {"_id": ObjectId(entry_id)}, {"$set": update}, return_document=ReturnDocument.AFTER
    )
    if not result:
        raise HTTPException(status_code=404, detail="Entry not found")
    await log_activity(user["username"], "dispatch_updated", result.get("dispatch_id", entry_id), "dispatch")
    return DispatchEntry.from_mongo(result).model_dump()


@router.delete("/{entry_id}")
async def delete_entry(entry_id: str, user: dict = Depends(get_current_user)):
    if not ObjectId.is_valid(entry_id):
        raise HTTPException(status_code=400, detail="Invalid entry ID")
    doc = await db.dispatch_entries.find_one_and_delete({"_id": ObjectId(entry_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Entry not found")
    await log_activity(user["username"], "dispatch_deleted", doc.get("dispatch_id", entry_id), "dispatch")
    return {"message": "Entry deleted"}


EXPORT_COLUMNS = [
    ("Dispatch ID", "dispatch_id"), ("Invoice No", "invoice_number"), ("Invoice Date", "invoice_date"),
    ("Customer", "customer_name"), ("Customer Code", "customer_code"), ("PO Number", "po_number"),
    ("Part No", "part_number"), ("Description", "part_description"), ("Qty", "quantity"),
    ("Unit", "unit"), ("Rate", "rate"), ("Total Value", "total_value"), ("GST", "gst"),
    ("Vehicle", "vehicle"), ("Dispatch Date", "dispatch_date"), ("Vendor", "vendor_name"),
]


@router.get("/export/excel")
async def export_excel(
    search: str = None, invoice: str = None, part: str = None, customer: str = None,
    date_from: str = None, date_to: str = None, user: dict = Depends(get_current_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    query = build_query(search, invoice, part, customer, date_from, date_to)
    docs = await db.dispatch_entries.find(query).sort("created_at", -1).to_list(5000)
    wb = Workbook()
    ws = wb.active
    ws.title = "Dispatch Entries"
    ws.append([c[0] for c in EXPORT_COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    for d in docs:
        ws.append([d.get(c[1], "") for c in EXPORT_COLUMNS])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    await log_activity(user["username"], "export_excel", f"{len(docs)} rows", "dispatch")
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=dispatch_entries.xlsx"},
    )


@router.get("/export/pdf")
async def export_pdf(
    search: str = None, invoice: str = None, part: str = None, customer: str = None,
    date_from: str = None, date_to: str = None, user: dict = Depends(get_current_user),
):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    query = build_query(search, invoice, part, customer, date_from, date_to)
    docs = await db.dispatch_entries.find(query).sort("created_at", -1).to_list(2000)
    cols = [("Dispatch ID", "dispatch_id"), ("Invoice No", "invoice_number"), ("Date", "invoice_date"),
            ("Customer", "customer_name"), ("Part No", "part_number"), ("Qty", "quantity"),
            ("Rate", "rate"), ("Total", "total_value"), ("Vendor", "vendor_name")]
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=15 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Grewal Engineering Works — Dispatch Report", styles["Title"]),
        Paragraph(f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} | Records: {len(docs)}", styles["Normal"]),
        Spacer(1, 8),
    ]
    data = [[c[0] for c in cols]] + [[str(d.get(c[1], ""))[:30] for c in cols] for d in docs]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F97316")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    await log_activity(user["username"], "export_pdf", f"{len(docs)} rows", "dispatch")
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=dispatch_report.pdf"},
    )
