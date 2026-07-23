import os
import io
import uuid
import logging
from typing import Optional
from datetime import datetime, timezone, timedelta
from fastapi import APIRouter, HTTPException, Depends, BackgroundTasks
from fastapi.responses import StreamingResponse, FileResponse
from bson import ObjectId
from pydantic import BaseModel
from database import db
from models import utcnow
from auth import get_current_user, require_admin, log_activity
from alerts import send_alert
from automation import (
    EWayBillAutomation, AutomationError, REQUIRED_ENV,
    load_selectors, save_selectors, validate_portal,
)
from pathlib import Path
from routes.worker_routes import (
    create_automation_job,
    desktop_execution_enabled,
    require_desktop_worker,
)

router = APIRouter(prefix="/eway", tags=["eway"])

@router.delete("/records/{record_id}")
async def delete_record(record_id: str, user: dict = Depends(require_admin)):
    if not ObjectId.is_valid(record_id):
        raise HTTPException(status_code=400, detail="Invalid E-Way record id")
    # The E-Way table is joined to master_dispatch; its row id is stored in
    # submissions.record_id, not submissions._id.  Deleting by _id made every
    # normal row look "missing" (and consequently undeletable).
    submission = await db.eway_submissions.find_one({"record_id": record_id})
    if submission and submission.get("status") == "Processing":
        raise HTTPException(status_code=409, detail="Processing E-Way records cannot be deleted")
    result = await db.eway_submissions.delete_one({"record_id": record_id})
    if not result.deleted_count and not await db.master_dispatch.find_one({"_id": ObjectId(record_id)}):
        raise HTTPException(status_code=404, detail="E-Way record not found")
    await log_activity(user["username"], "eway_record_deleted", record_id, "eway")
    return {"deleted": True}
logger = logging.getLogger(__name__)

ROOT_DIR = Path(__file__).parent.parent
MAX_RETRIES = 3

run_state = {"running": False, "run_id": None, "module": None, "total": 0, "processed": 0, "started_at": None}


class RunRequest(BaseModel):
    ids: list[str] = []
    dry_run: bool = False


class ModeRequest(BaseModel):
    mode: str


class ValidateRequest(BaseModel):
    attempt_login: bool = False
    dry_run_fill: bool = False


class EwayDetailsInput(BaseModel):
    company_code: str = "TMTL"
    from_validity: str = ""
    to_validity: str = ""


def now_iso():
    return utcnow().isoformat()


async def get_mode():
    from environment import get_effective_automation_mode
    return await get_effective_automation_mode()


def make_logger(run_id, module):
    async def log(event, message, dispatch_id=None, level="INFO"):
        await db.automation_logs.insert_one({
            "id": str(uuid.uuid4()), "run_id": run_id, "module": module,
            "event": event, "message": message, "dispatch_id": dispatch_id,
            "level": level, "timestamp": now_iso(),
        })
    return log


async def get_submission(record_id: str) -> dict:
    return await db.eway_submissions.find_one({"record_id": record_id}) or {}


def join_record(md: dict, sub: dict) -> dict:
    return {
        "id": str(md["_id"]),
        "dispatch_no": md.get("dispatch_no", ""),
        "invoice_no": md.get("invoice_number", ""),
        "invoice_date": md.get("invoice_date", ""),
        "customer_name": md.get("customer_name", ""),
        "eway_bill_number": md.get("eway_bill_number", "") or "",
        "company_code": sub.get("company_code", "TMTL"),
        "from_validity": sub.get("from_validity", ""),
        "to_validity": sub.get("to_validity", ""),
        "eway_status": sub.get("status", "Pending"),
        "completed_time": sub.get("completed_time"),
        "error": sub.get("error"),
        "retry_count": sub.get("retry_count", 0),
        "submitted_by": sub.get("submitted_by"),
        "screenshot": sub.get("screenshot"),
    }


async def build_record_query(status=None, invoice=None, dispatch=None, date=None):
    from environment import env_list_filter
    query = await env_list_filter()
    if invoice:
        query["invoice_number"] = {"$regex": invoice, "$options": "i"}
    if dispatch:
        query["dispatch_no"] = {"$regex": dispatch, "$options": "i"}
    if date:
        query["invoice_date"] = date
    if status and status != "All":
        if status == "Pending":
            done_ids = [ObjectId(s["record_id"]) async for s in db.eway_submissions.find(
                {"status": {"$in": ["Completed", "Failed"]}}, {"record_id": 1}) if ObjectId.is_valid(s["record_id"])]
            query["_id"] = {"$nin": done_ids}
        else:
            ids = [ObjectId(s["record_id"]) async for s in db.eway_submissions.find(
                {"status": status}, {"record_id": 1}) if ObjectId.is_valid(s["record_id"])]
            query["_id"] = {"$in": ids}
    return query


# ---------- Batch runner on the shared automation engine ----------

def format_eway(v: str) -> str:
    """Portal format: XXXX XXXX XXXX (12 digits in groups of 4)."""
    import re as _re
    digits = _re.sub(r"\D", "", v or "")
    if len(digits) == 12:
        return f"{digits[0:4]} {digits[4:8]} {digits[8:12]}"
    return (v or "").strip()


def eway_prepare(md: dict, sub: dict):
    if not (md.get("eway_bill_number") or "").strip():
        return None, "E-Way Bill Number is blank"
    if not (sub.get("from_validity") or "").strip() or not (sub.get("to_validity") or "").strip():
        return None, "validity dates are blank"
    return {
        "company_code": (sub.get("company_code") or "").strip() or "TMTL",
        "eway_bill_number": format_eway(md["eway_bill_number"]),
        "eway_from_validity": sub["from_validity"],
        "eway_to_validity": sub["to_validity"],
    }, None


async def set_submission(record_id: str, fields: dict, inc: dict = None):
    update = {"$set": {**fields, "record_id": record_id, "updated_at": now_iso()}}
    if inc:
        update["$inc"] = inc
    await db.eway_submissions.update_one({"record_id": record_id}, update, upsert=True)


async def process_batch(ids: list[str], run_id: str, user: str, force_mode: str = None):
    mode = force_mode or await get_mode()
    headless = os.environ.get("AUTOMATION_HEADLESS", "true").lower() == "true"
    log = make_logger(run_id, "eway")
    bot = EWayBillAutomation(mode=mode, headless=headless, log=log)
    try:
        await log("Run Started", f"E-Way Bill Entry: processing {len(ids)} record(s) in {mode.upper()} mode (user: {user})")
        if mode == "live":
            bot.require_env()
        await bot.start()
        await bot.login()
        await bot.navigate_to_entry()
        for rec_id in ids:
            md = await db.master_dispatch.find_one({"_id": ObjectId(rec_id)}) if ObjectId.is_valid(rec_id) else None
            if not md:
                run_state["processed"] += 1
                continue
            d_id = md.get("dispatch_no", rec_id)
            sub = await get_submission(rec_id)
            data, skip_reason = eway_prepare(md, sub)
            if skip_reason:
                await log("Skipped", f"Dispatch {d_id}: {skip_reason} - record skipped", d_id, "WARN")
                await set_submission(rec_id, {"error": f"Skipped: {skip_reason}"})
                run_state["processed"] += 1
                continue
            await log("Record Started", f"Dispatch {d_id}: filling E-Way Bill Entry form", d_id)
            attempts, success, last_err = 0, False, None
            while attempts < MAX_RETRIES and not success:
                attempts += 1
                try:
                    await bot.fill_form(data)
                    await bot.submit()
                    await log("Record Submitted", f"Dispatch {d_id}: form submitted (attempt {attempts})", d_id)
                    await bot.verify_success()
                    success = True
                except AutomationError as e:
                    last_err = str(e)
                    if attempts < MAX_RETRIES:
                        await log("Retry", f"Dispatch {d_id}: attempt {attempts} failed ({last_err}) - retrying", d_id, "WARN")
                    else:
                        await log("Error", f"Dispatch {d_id}: failed after {attempts} attempts - {last_err}", d_id, "ERROR")
            if success:
                ts = now_iso()
                await set_submission(rec_id, {
                    "status": "Completed", "completed_time": ts, "error": None,
                    "retry_count": 0, "submitted_by": user, "dispatch_no": d_id,
                })
                await db.master_dispatch.update_one(
                    {"_id": md["_id"]}, {"$set": {"status": "completed", "updated_at": ts}}
                )
                await log("Record Completed", f"Dispatch {d_id}: E-Way Bill Entry completed", d_id, "SUCCESS")
            else:
                shot = await bot.capture_screenshot(f"eway_fail_{d_id}")
                await set_submission(rec_id, {
                    "status": "Failed", "error": last_err, "screenshot": shot, "dispatch_no": d_id,
                }, inc={"retry_count": 1})
                await send_alert("E-Way Bill automation failed",
                                 f"Dispatch {d_id}: {str(last_err)[:300]}")
            run_state["processed"] += 1
    except AutomationError as e:
        await log("Error", str(e), level="ERROR")
    except Exception as e:
        logger.exception("E-Way automation run failed")
        await log("Error", f"Unexpected Error: {type(e).__name__}", level="ERROR")
    finally:
        await bot.close()
        await log("Finished", f"Run finished: {run_state['processed']}/{run_state['total']} record(s) processed")
        run_state["running"] = False


async def start_run(background_tasks: BackgroundTasks, ids: list[str], user: str):
    mode = await get_mode()  # blocks in maintenance / emergency stop before scheduling
    if run_state["running"]:
        raise HTTPException(status_code=409, detail="An automation run is already in progress")
    if not ids:
        raise HTTPException(status_code=400, detail="No records to process")

    if desktop_execution_enabled():
        is_test = mode != "live"
        await require_desktop_worker("eway_bill_entry", allow_offline_test=True, test_mode=is_test)
        jobs, skipped = [], []
        for rec_id in ids:
            md = await db.master_dispatch.find_one({"_id": ObjectId(rec_id)}) if ObjectId.is_valid(rec_id) else None
            if not md:
                continue
            sub = await get_submission(rec_id)
            payload, skip_reason = eway_prepare(md, sub)
            if skip_reason:
                skipped.append({"record_id": rec_id, "reason": skip_reason})
                await set_submission(rec_id, {"error": f"Skipped: {skip_reason}"})
                continue
            job = await create_automation_job(
                job_type="eway_bill_entry", payload={**payload, "dry_run": req.dry_run}, source_record_id=rec_id,
                created_by=user, test_mode=is_test, priority=70,
            )
            await set_submission(rec_id, {
                "status": "Queued", "desktop_job_id": job["id"], "error": None,
                "submitted_by": user, "dispatch_no": md.get("dispatch_no", rec_id),
            })
            jobs.append(job)
        return {"execution": "desktop", "queued": len(jobs), "jobs": jobs,
                "skipped": skipped, "module": "eway"}

    run_id = str(uuid.uuid4())
    run_state.update({"running": True, "run_id": run_id, "module": "eway",
                      "total": len(ids), "processed": 0, "started_at": now_iso()})
    background_tasks.add_task(process_batch, ids, run_id, user)
    return {"run_id": run_id, "total": len(ids), "module": "eway"}


async def ids_by_status(status: str) -> list[str]:
    query = await build_record_query(status=status)
    docs = await db.master_dispatch.find(query, {"_id": 1}).to_list(2000)
    return [str(d["_id"]) for d in docs]


# ---------- Records ----------

@router.get("/records")
async def eway_records(status: Optional[str] = None, invoice: Optional[str] = None,
                       dispatch: Optional[str] = None, date: Optional[str] = None,
                       page: int = 1, page_size: int = 50, user: dict = Depends(get_current_user)):
    query = await build_record_query(status, invoice, dispatch, date)
    page = max(1, page)
    page_size = min(max(1, page_size), 200)
    total = await db.master_dispatch.count_documents(query)
    docs = await db.master_dispatch.find(query).sort("created_at", -1).skip((page - 1) * page_size).to_list(page_size)
    ids = [str(d["_id"]) for d in docs]
    subs = {s["record_id"]: s async for s in db.eway_submissions.find({"record_id": {"$in": ids}})}
    return {"items": [join_record(d, subs.get(str(d["_id"]), {})) for d in docs],
            "total": total, "page": page, "pages": max(1, -(-total // page_size))}


@router.put("/records/{record_id}")
async def update_eway_details(record_id: str, body: EwayDetailsInput, user: dict = Depends(get_current_user)):
    if not ObjectId.is_valid(record_id):
        raise HTTPException(status_code=400, detail="Invalid record ID")
    md = await db.master_dispatch.find_one({"_id": ObjectId(record_id)})
    if not md:
        raise HTTPException(status_code=404, detail="Master Dispatch record not found")
    await set_submission(record_id, {
        "company_code": body.company_code.strip() or "TMTL",
        "from_validity": body.from_validity.strip(),
        "to_validity": body.to_validity.strip(),
        "dispatch_no": md.get("dispatch_no", ""),
    })
    await log_activity(user["username"], "eway_details_updated", md.get("dispatch_no", record_id), "eway")
    return join_record(md, await get_submission(record_id))


@router.get("/stats")
async def eway_stats(user: dict = Depends(get_current_user)):
    total = await db.master_dispatch.count_documents({})
    completed = await db.eway_submissions.count_documents({"status": "Completed"})
    failed = await db.eway_submissions.count_documents({"status": "Failed"})
    return {"total": total, "pending": max(0, total - completed - failed),
            "completed": completed, "failed": failed}


# ---------- Runs ----------

@router.post("/run")
async def eway_run(req: RunRequest, background_tasks: BackgroundTasks, user: dict = Depends(get_current_user)):
    result = await start_run(background_tasks, req.ids, user["username"])
    await log_activity(user["username"], "eway_run", f"{len(req.ids)} record(s)", "eway")
    return result


@router.post("/run-all-pending")
async def eway_run_all(background_tasks: BackgroundTasks, user: dict = Depends(get_current_user)):
    ids = await ids_by_status("Pending")
    result = await start_run(background_tasks, ids, user["username"])
    await log_activity(user["username"], "eway_run_all_pending", f"{len(ids)} record(s)", "eway")
    return result


@router.post("/retry-failed")
async def eway_retry_failed(background_tasks: BackgroundTasks, user: dict = Depends(get_current_user)):
    ids = await ids_by_status("Failed")
    result = await start_run(background_tasks, ids, user["username"])
    await log_activity(user["username"], "eway_retry_failed", f"{len(ids)} record(s)", "eway")
    return result


@router.get("/run-status")
async def eway_run_status(user: dict = Depends(get_current_user)):
    return run_state


@router.get("/logs")
async def eway_logs(run_id: Optional[str] = None, limit: int = 150, user: dict = Depends(get_current_user)):
    q = {"module": {"$in": ["eway", "portal_validation"]}}
    if run_id:
        q["run_id"] = run_id
    logs = await db.automation_logs.find(q, {"_id": 0}).sort("timestamp", -1).to_list(min(limit, 500))
    return list(reversed(logs))


# ---------- Export ----------

@router.get("/export")
async def eway_export(status: Optional[str] = None, invoice: Optional[str] = None,
                      dispatch: Optional[str] = None, date: Optional[str] = None,
                      user: dict = Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    query = await build_record_query(status, invoice, dispatch, date)
    docs = await db.master_dispatch.find(query).sort("created_at", -1).to_list(5000)
    ids = [str(d["_id"]) for d in docs]
    subs = {s["record_id"]: s async for s in db.eway_submissions.find({"record_id": {"$in": ids}})}
    wb = Workbook()
    ws = wb.active
    ws.title = "E-Way Bill Entry"
    ws.append(["Dispatch No", "Invoice No", "Invoice Date", "Customer", "Company Code", "E-Way Bill No",
               "From Validity", "To Validity", "Status", "Completed Time", "Error", "Retry Count", "Submitted By"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    for d in docs:
        r = join_record(d, subs.get(str(d["_id"]), {}))
        ws.append([r["dispatch_no"], r["invoice_no"], r["invoice_date"], r["customer_name"],
                   r["company_code"], r["eway_bill_number"], r["from_validity"], r["to_validity"],
                   r["eway_status"], r["completed_time"], r["error"], r["retry_count"], r["submitted_by"]])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    await log_activity(user["username"], "eway_export", f"{len(docs)} rows", "eway")
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=eway_bill_entries.xlsx"})


@router.get("/screenshots/{name}")
async def eway_screenshot(name: str, user: dict = Depends(get_current_user)):
    if "/" in name or ".." in name:
        raise HTTPException(status_code=400, detail="Invalid screenshot name")
    path = ROOT_DIR / "screenshots" / name
    if not path.exists():
        raise HTTPException(status_code=404, detail="Screenshot not found")
    return FileResponse(str(path), media_type="image/png")


# ---------- Settings, selectors & validation ----------

@router.get("/settings")
async def eway_settings(user: dict = Depends(get_current_user)):
    from environment import get_environment
    env = await get_environment()
    mode = env["mode"] if env["mode"] in ("test", "live") else "test"
    missing = [v for v in REQUIRED_ENV if not os.environ.get(v)]
    return {"mode": mode, "env_configured": not missing, "missing_env": missing,
            "headless": os.environ.get("AUTOMATION_HEADLESS", "true").lower() == "true"}


@router.post("/settings/mode")
async def eway_set_mode(req: ModeRequest, user: dict = Depends(require_admin)):
    raise HTTPException(status_code=400,
                        detail="Automation mode is controlled centrally — use Settings → System Environment")


@router.get("/selectors")
async def eway_get_selectors(user: dict = Depends(get_current_user)):
    return load_selectors()


@router.put("/selectors")
async def eway_update_selectors(payload: dict, user: dict = Depends(require_admin)):
    required_sections = ("login", "eway")
    missing = [s for s in required_sections if s not in payload]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing selector section(s): {', '.join(missing)}")
    import json as _json
    changed = _json.dumps(load_selectors(), sort_keys=True) != _json.dumps(payload, sort_keys=True)
    if changed:
        await db.settings.delete_many({"key": {"$in": ["last_portal_validation", "last_test_validation"]}})
    save_selectors(payload)
    await log_activity(user["username"], "eway_selector_update",
                       "Selectors changed — validations reset" if changed else "No changes", "eway")
    return {"ok": True, "changed": bool(changed)}


@router.post("/portal/validate")
async def eway_portal_validate(req: ValidateRequest, user: dict = Depends(require_admin)):
    if run_state["running"]:
        raise HTTPException(status_code=409, detail="An automation run is already in progress")
    if desktop_execution_enabled():
        mode = await get_mode()
        is_test = mode != "live"
        await require_desktop_worker("portal_validation", allow_offline_test=False, test_mode=is_test)
        job = await create_automation_job(
            job_type="portal_validation",
            payload={"attempt_login": req.attempt_login, "dry_run_fill": req.dry_run_fill},
            source_record_id="portal_validation", created_by=user["username"],
            test_mode=False, priority=10,
        )
        return {"queued": True, "execution": "desktop", "job": job, "results": [],
                "passed": 0, "total": 0, "all_ok": False, "full_validation": False}

    headless = os.environ.get("AUTOMATION_HEADLESS", "true").lower() == "true"
    log = make_logger(str(uuid.uuid4()), "portal_validation")
    results = await validate_portal(attempt_login=req.attempt_login, headless=headless, log=log,
                                    dry_run_fill=req.dry_run_fill)
    passed = sum(1 for r in results if r["status"] == "ok")
    all_ok = passed == len(results) and len(results) > 0
    full_ok = all_ok and req.attempt_login
    failed_steps = [r["step"] for r in results if r["status"] != "ok"]
    await db.settings.update_one({"key": "last_portal_validation"}, {"$set": {"value": {
        "timestamp": now_iso(), "attempt_login": req.attempt_login,
        "passed": passed, "total": len(results), "all_ok": full_ok, "failed_steps": failed_steps,
    }}}, upsert=True)
    await log_activity(user["username"], "eway_portal_validation", f"{passed}/{len(results)} checks passed", "eway")
    return {"results": results, "passed": passed, "total": len(results),
            "all_ok": all_ok, "full_validation": full_ok}


@router.get("/validation/status")
async def eway_validation_status(user: dict = Depends(get_current_user)):
    pv = await db.settings.find_one({"key": "last_portal_validation"}, {"_id": 0})
    tv = await db.settings.find_one({"key": "last_test_validation"}, {"_id": 0})
    portal = pv["value"] if pv else None
    test_run = tv["value"] if tv else None
    from environment import get_environment
    env = await get_environment()
    return {"portal_validation": portal, "test_validation": test_run,
            "ready_for_live": bool(portal and portal.get("all_ok") and test_run and test_run.get("all_ok")),
            "mode": env["mode"] if env["mode"] in ("test", "live") else "test"}


@router.post("/validation/test-run")
async def eway_validation_test_run(user: dict = Depends(require_admin)):
    """Complete TEST-mode end-to-end workflow validation on temporary sample records."""
    if run_state["running"]:
        raise HTTPException(status_code=409, detail="An automation run is already in progress")
    run_id = str(uuid.uuid4())
    suffix = run_id[:8].upper()
    today = datetime.now(timezone.utc)
    fmt = "%d/%m/%Y"
    samples = [
        (f"VAL-GOOD-{suffix}", "351099999901", today.strftime(fmt), (today + timedelta(days=5)).strftime(fmt)),
        (f"VAL-FAIL-{suffix}", f"ERR-{suffix}", today.strftime(fmt), (today + timedelta(days=3)).strftime(fmt)),
        (f"VAL-SKIP-{suffix}", "", "", ""),
    ]
    ids = []
    for d_id, bill, from_v, to_v in samples:
        result = await db.master_dispatch.insert_one({
            "dispatch_no": d_id, "invoice_number": f"INV-{d_id}", "eway_bill_number": bill,
            "status": "pending", "verified": True, "items": [], "created_by": user["username"],
            "created_at": now_iso(), "updated_at": now_iso(),
        })
        rid = str(result.inserted_id)
        if from_v:
            await set_submission(rid, {"company_code": "TMTL", "from_validity": from_v, "to_validity": to_v})
        ids.append(rid)
    run_state.update({"running": True, "run_id": run_id, "module": "test_validation",
                      "total": len(ids), "processed": 0, "started_at": now_iso()})
    checks = []
    try:
        await process_batch(ids, run_id, user["username"], force_mode="test")
        good = await get_submission(ids[0])
        fail = await get_submission(ids[1])
        skip = await get_submission(ids[2])
        good_md = await db.master_dispatch.find_one({"_id": ObjectId(ids[0])})
        logs = await db.automation_logs.find({"run_id": run_id}, {"_id": 0}).to_list(500)
        events = [l["event"] for l in logs]

        def check(name, ok, detail):
            checks.append({"check": name, "status": "ok" if ok else "fail", "detail": detail})

        check("Login", "Login Success" in events, "Engine logged in to portal (simulated in TEST mode)")
        check("Navigation", "Navigation" in events, "Navigated to E-Way Bill -> E-Way Bill Entry")
        check("Form entry & submission", "Record Submitted" in events, "Form filled, verified and submitted")
        check("Success detection", good.get("status") == "Completed" and bool(good.get("completed_time")),
              f"VAL-GOOD-{suffix} confirmed and marked Completed")
        check("Master Dispatch sync", bool(good_md) and good_md.get("status") == "completed",
              "Master Dispatch record status updated to 'completed' on success")
        check("Retry logic", events.count("Retry") >= 2,
              f"VAL-FAIL-{suffix} retried {events.count('Retry')} time(s) before failing (max 3 attempts)")
        check("Failure handling", fail.get("status") == "Failed" and fail.get("retry_count", 0) >= 1 and bool(fail.get("error")),
              f"VAL-FAIL-{suffix} marked Failed, error stored: '{fail.get('error')}'")
        shot = fail.get("screenshot")
        shot_ok = bool(shot) and (ROOT_DIR / shot).exists()
        check("Screenshot capture on failure", shot_ok, f"Screenshot saved: {shot}" if shot_ok else "No screenshot found")
        check("Validation skip rules", "Skipped" in events and (skip.get("error") or "").startswith("Skipped"),
              f"VAL-SKIP-{suffix} with blank E-Way Bill correctly skipped")
        check("Database updates", good.get("retry_count") == 0 and good.get("error") is None and good.get("submitted_by") == user["username"],
              "Status, timestamps, retry count, user and error fields updated correctly")
        check("Logging", all(e in events for e in ("Run Started", "Record Started", "Record Completed", "Finished")),
              "All lifecycle events present in automation log")
    finally:
        await db.master_dispatch.delete_many({"_id": {"$in": [ObjectId(i) for i in ids]}})
        await db.eway_submissions.delete_many({"record_id": {"$in": ids}})
        run_state["running"] = False
    all_ok = all(c["status"] == "ok" for c in checks)
    summary = {"timestamp": now_iso(), "all_ok": all_ok,
               "passed": sum(1 for c in checks if c["status"] == "ok"), "total": len(checks)}
    await db.settings.update_one({"key": "last_test_validation"}, {"$set": {"value": summary}}, upsert=True)
    await log_activity(user["username"], "eway_test_validation", f"{summary['passed']}/{summary['total']} checks passed", "eway")
    return {"checks": checks, "all_ok": all_ok, **summary}
