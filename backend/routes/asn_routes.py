import os
import io
import re
import uuid
import time
import asyncio
import logging
from typing import Optional
from datetime import datetime, timezone
from fastapi import APIRouter, HTTPException, Depends, BackgroundTasks, UploadFile, File
from fastapi.responses import StreamingResponse, FileResponse
from bson import ObjectId
from pydantic import BaseModel, Field
from pathlib import Path
from database import db
from environment import env_fields, env_list_filter
from models import utcnow
from auth import get_current_user, require_admin, log_activity
from alerts import send_alert
from automation import ASNAutomation, AutomationError, AsnValidationError, DropdownMatchError, BatchAllocationError, SCREENSHOT_DIR
from routes.worker_routes import (
    create_automation_job,
    desktop_execution_enabled,
    require_desktop_worker,
)

router = APIRouter(prefix="/asn", tags=["asn"])

@router.delete("/records/{record_id}")
async def delete_record(record_id: str, user: dict = Depends(require_admin)):
    if not ObjectId.is_valid(record_id):
        raise HTTPException(status_code=400, detail="Invalid ASN record id")
    result = await db.asn_creation.delete_one({"_id": ObjectId(record_id), "status": {"$nin": ["Processing"]}})
    if not result.deleted_count:
        raise HTTPException(status_code=409, detail="Processing or missing ASN records cannot be deleted")
    await db.asn_batch_allocations.delete_many({"asn_record_id": record_id})
    return {"deleted": True}
logger = logging.getLogger(__name__)

ROOT_DIR = Path(__file__).parent.parent
PDI_DIR = ROOT_DIR / "uploads" / "pdi"
PDI_DIR.mkdir(parents=True, exist_ok=True)
MAX_RETRIES = 3
ALLOCATION_TIMEOUT = 900

run_state = {"running": False, "run_id": None, "total": 0, "processed": 0, "current": None,
             "started_at": None, "awaiting_allocation": None}
alloc_state = {"event": None, "record_id": None, "batches": [], "asn_qty": 0, "result": None, "cancelled": False}


class RunRequest(BaseModel):
    ids: list[str] = []


class AsnEditInput(BaseModel):
    po_number: str = ""
    transporter: str = ""
    basic_amount: float = Field(0, ge=0)
    total_amount: float = Field(0, ge=0)


def now_iso():
    return utcnow().isoformat()


async def get_mode():
    from environment import get_effective_automation_mode
    return await get_effective_automation_mode()


def serialize(a: dict) -> dict:
    a = dict(a)
    a["id"] = str(a.pop("_id"))
    return a


def compute_status(doc: dict) -> str:
    ready = all(str(doc.get(k) or "").strip() for k in ("po_number", "invoice_no", "invoice_date", "transporter")) \
        and doc.get("items") and float(doc.get("total_amount") or 0) > 0 \
        and str(doc.get("pdi_file_path") or "").strip()
    return "Ready" if ready else "Draft"


def to_dmy(iso: str) -> str:
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", iso or "")
    return f"{m.group(3)}/{m.group(2)}/{m.group(1)}" if m else (iso or "")


async def append_log(oid, run_id, event, message, level="INFO"):
    entry = {"ts": now_iso(), "event": event, "message": message, "level": level}
    await db.asn_creation.update_one({"_id": oid}, {"$push": {"automation_log": entry}, "$set": {"updated_at": now_iso()}})
    await db.automation_logs.insert_one({"id": str(uuid.uuid4()), "run_id": run_id, "module": "asn",
                                         "event": event, "message": message, "level": level, "timestamp": now_iso()})


async def process_queue(ids: list[str], run_id: str, user: str):
    """Queue processing: one ASN at a time on a single browser session."""
    mode = await get_mode()
    headless = os.environ.get("AUTOMATION_HEADLESS", "true").lower() == "true"
    current_oid = {"v": None}

    async def log(event, message, dispatch_id=None, level="INFO"):
        if current_oid["v"] is not None:
            await append_log(current_oid["v"], run_id, event, message, level)

    record_allocs = {"v": []}

    async def request_allocation(part, asn_qty, batches):
        """Pauses the automation until the user confirms/cancels a batch allocation in the UI."""
        oid = current_oid["v"]
        rid = str(oid)
        prior = [a for a in record_allocs["v"] if a["part_number"] == part]
        if prior and abs(sum(float(a["allocated_quantity"]) for a in prior) - float(asn_qty)) < 0.001:
            avail = {b["batch_no"]: float(b["available_qty"]) for b in batches}
            if all(a["batch_number"] in avail and float(a["allocated_quantity"]) <= avail[a["batch_number"]] for a in prior):
                await append_log(oid, run_id, "Batch Allocation", f"Reusing previously confirmed allocation for part {part}")
                return [{"batch_no": a["batch_number"], "allocate_qty": float(a["allocated_quantity"]),
                         "consider": a["batch_considerable"] == "Yes"} for a in prior]
        ev = asyncio.Event()
        alloc_state.update({"event": ev, "record_id": rid, "batches": batches,
                            "asn_qty": float(asn_qty), "result": None, "cancelled": False})
        run_state["awaiting_allocation"] = {"record_id": rid, "invoice_no": run_state["current"],
                                            "part_number": part, "asn_qty": float(asn_qty), "batches": batches}
        await db.asn_creation.update_one({"_id": oid}, {"$set": {"status": "Awaiting Allocation", "updated_at": now_iso()}})
        await append_log(oid, run_id, "Batch Allocation", f"Waiting for user allocation - part {part}, ASN Qty {asn_qty}, {len(batches)} batch(es)")
        try:
            await asyncio.wait_for(ev.wait(), timeout=ALLOCATION_TIMEOUT)
        except asyncio.TimeoutError:
            raise BatchAllocationError("Batch allocation timed out (no user response within 15 minutes)")
        finally:
            run_state["awaiting_allocation"] = None
            alloc_state["event"] = None
        if alloc_state["cancelled"] or not alloc_state["result"]:
            raise BatchAllocationError("Batch allocation cancelled by user")
        await db.asn_creation.update_one({"_id": oid}, {"$set": {"status": "Processing", "updated_at": now_iso()}})
        result = alloc_state["result"]
        record_allocs["v"] = [a for a in record_allocs["v"] if a["part_number"] != part]
        by_no = {b["batch_no"]: b for b in batches}
        for a in result:
            b = by_no.get(a["batch_no"], {})
            record_allocs["v"].append({
                "part_number": part, "batch_number": a["batch_no"],
                "batch_quantity": float(b.get("batch_qty") or 0), "available_quantity": float(b.get("available_qty") or 0),
                "allocated_quantity": float(a["allocate_qty"]), "batch_considerable": "Yes" if a["consider"] else "No",
            })
        return result

    bot = ASNAutomation(mode=mode, headless=headless, log=log)
    bot.allocation_cb = request_allocation
    try:
        if mode == "live":
            bot.require_env()
        await bot.start()
        await bot.login()
        for rec_id in ids:
            oid = ObjectId(rec_id)
            current_oid["v"] = oid
            doc = await db.asn_creation.find_one({"_id": oid})
            if not doc or doc.get("status") == "Completed":
                run_state["processed"] += 1
                continue
            run_state["current"] = doc.get("invoice_no")
            record_allocs["v"] = list(doc.get("batch_allocations") or [])
            await db.asn_creation.update_one({"_id": oid}, {"$set": {"status": "Processing", "error_message": "", "updated_at": now_iso()}})
            data = {
                "po_number": doc.get("po_number", ""), "invoice_no": doc.get("invoice_no", ""),
                "invoice_date": to_dmy(doc.get("invoice_date", "")),
                "basic_amount": doc.get("basic_amount", 0), "total_amount": doc.get("total_amount", 0),
                "cgst": doc.get("cgst", 0), "sgst": doc.get("sgst", 0), "igst": doc.get("igst", 0),
                "no_of_cases": doc.get("no_of_cases", 0),
                "transporter": doc.get("transporter", ""), "items": doc.get("items", []),
                "pdi_path": doc.get("pdi_file_path", ""),
            }
            status, error_message, asn_number, screenshots = "Failed", "", "", dict(doc.get("screenshots") or {})
            attempts = 0
            try:
                await log("Run Started", f"Creating ASN for invoice {doc.get('invoice_no')} in {mode.upper()} mode by {user}")
                await bot.navigate_to_entry()
                while True:
                    try:
                        result = await bot.run_asn(data)
                        status, asn_number = "Completed", result["asn_number"]
                        if result.get("before_submit"):
                            screenshots["before_submit"] = result["before_submit"]
                        shot = await bot.capture_screenshot(f"asn_success_{rec_id[-6:]}")
                        if shot:
                            screenshots["after_success"] = shot
                        break
                    except (AsnValidationError, DropdownMatchError, BatchAllocationError) as e:
                        error_message = str(e)
                        await log("Error", error_message, level="ERROR")
                        break
                    except AutomationError as e:
                        attempts += 1
                        if attempts < MAX_RETRIES:
                            await log("Retry", f"Attempt {attempts} failed ({e}) - retrying", level="WARN")
                            continue
                        error_message = str(e)
                        await log("Error", f"Failed after {attempts} attempts: {e}", level="ERROR")
                        break
            except Exception as e:
                logger.exception("ASN automation failed")
                error_message = f"Unexpected Error: {type(e).__name__}: {str(e)[:200]}"
                await log("Error", error_message, level="ERROR")
            if status == "Failed":
                shot = await bot.capture_screenshot(f"asn_fail_{rec_id[-6:]}")
                if shot:
                    screenshots["after_failure"] = shot
                page_url, html_file = "", ""
                if bot.page:
                    try:
                        page_url = bot.page.url
                        html_name = f"asn_fail_{rec_id[-6:]}_{uuid.uuid4().hex[:6]}.html"
                        (SCREENSHOT_DIR / html_name).write_text(await bot.page.content())
                        html_file = f"screenshots/{html_name}"
                    except Exception:
                        pass
                await db.asn_creation.update_one({"_id": oid}, {"$set": {"failure_url": page_url, "failure_html": html_file}})
            update = {"status": status, "error_message": error_message, "screenshots": screenshots, "updated_at": now_iso()}
            if record_allocs["v"]:
                update["batch_allocations"] = record_allocs["v"]
            if status == "Completed":
                update["asn_number"] = asn_number
                update["completed_at"] = now_iso()
                if record_allocs["v"]:
                    await db.asn_batch_allocations.delete_many({"asn_record_id": rec_id})
                    await db.asn_batch_allocations.insert_many([{
                        "asn_record_id": rec_id, "asn_number": asn_number,
                        "dispatch_id": doc.get("master_dispatch_id", ""), "dispatch_no": doc.get("dispatch_no", ""),
                        "invoice_no": doc.get("invoice_no", ""), **a,
                        "created_by": user, "created_at": now_iso(),
                    } for a in record_allocs["v"]])
                    await log("Batch Allocation", f"{len(record_allocs['v'])} batch allocation(s) saved against {asn_number}", level="SUCCESS")
                if ObjectId.is_valid(doc.get("master_dispatch_id", "")):
                    await db.master_dispatch.update_one(
                        {"_id": ObjectId(doc["master_dispatch_id"])},
                        {"$set": {"asn_number": asn_number, "status": "ready_for_eway", "updated_at": now_iso()}},
                    )
                    await db.master_dispatch.update_one(
                        {"_id": ObjectId(doc["master_dispatch_id"]), "documents.type": "PDI"},
                        {"$set": {"documents.$.upload_status": "Uploaded to Portal",
                                  "documents.$.last_upload_at": now_iso(),
                                  "pdi_upload_status": "Uploaded to Portal",
                                  "pdi_last_upload_at": now_iso()}},
                    )
                    await log("ASN Number Captured", f"{asn_number} linked to Master Dispatch - available to E-Way Bill & Vendor Ack modules", level="SUCCESS")
            await db.asn_creation.update_one({"_id": oid}, {"$set": update})
            if status == "Failed":
                await send_alert("ASN automation failed",
                                 f"Invoice {doc.get('invoice_no')}: {str(error_message)[:300]}")
            run_state["processed"] += 1
    except AutomationError as e:
        logger.error(f"ASN queue aborted: {e}")
        await db.asn_creation.update_many(
            {"_id": {"$in": [ObjectId(i) for i in ids]}, "status": "Processing"},
            {"$set": {"status": "Failed", "error_message": str(e), "updated_at": now_iso()}},
        )
    finally:
        await bot.close()
        run_state["running"] = False
        run_state["current"] = None
        run_state["awaiting_allocation"] = None
        alloc_state["event"] = None


# ---------- Import ----------

@router.post("/import")
async def import_from_md(user: dict = Depends(get_current_user)):
    existing_ids = {d["master_dispatch_id"] async for d in db.asn_creation.find({}, {"master_dispatch_id": 1})}
    query = {"$or": [{"asn_number": ""}, {"asn_number": {"$exists": False}}], **(await env_list_filter())}
    imported = 0
    async for md in db.master_dispatch.find(query):
        mid = str(md["_id"])
        if mid in existing_ids:
            continue
        pdi_doc = next((d for d in (md.get("documents") or []) if d.get("type") == "PDI"), None)
        pdi_path = pdi_doc.get("file_path", "") if pdi_doc else ""
        doc = {
            "master_dispatch_id": mid, "dispatch_no": md.get("dispatch_no", ""),
            "invoice_no": md.get("invoice_number", ""), "invoice_date": md.get("invoice_date", ""),
            "po_number": md.get("po_number", ""), "supplier_name": "GREWAL ENGINEERING WORKS",
            "plant": md.get("plant", ""), "vehicle_number": md.get("vehicle_number", ""),
            "transporter": md.get("transporter_name", ""),
            "basic_amount": round(float(md.get("invoice_total") or 0) - float(md.get("gst_total") or 0), 2),
            "total_amount": float(md.get("invoice_total") or 0),
            "cgst": float(md.get("cgst") or 0), "sgst": float(md.get("sgst") or 0), "igst": float(md.get("igst") or 0),
            "no_of_cases": int(md.get("boxes") or 0),
            "items": [{"part_number": i.get("part_number", ""), "description": i.get("description", ""),
                       "quantity": i.get("quantity", 0)} for i in (md.get("items") or []) if i.get("part_number")],
            "pdi_file_path": pdi_path if pdi_path and os.path.exists(pdi_path) else "",
            "pdi_file_name": f"{pdi_doc.get('report_no', '')}.pdf" if pdi_doc and pdi_path and os.path.exists(pdi_path) else "",
            "asn_number": "", "error_message": "",
            "automation_log": [], "screenshots": {},
            "created_at": now_iso(), "updated_at": now_iso(), "created_by": user["username"],
        }
        doc["status"] = compute_status(doc)
        await db.asn_creation.insert_one({**doc, **(await env_fields())})
        imported += 1
    await log_activity(user["username"], "asn_import", f"{imported} record(s) imported from Master Dispatch", "asn")
    return {"imported": imported}


# ---------- Records ----------

@router.get("/records")
async def asn_records(status: Optional[str] = None, search: Optional[str] = None,
                      page: int = 1, page_size: int = 50, user: dict = Depends(get_current_user)):
    query = await env_list_filter()
    if status and status != "All":
        query["status"] = status
    if search:
        rx = {"$regex": re.escape(search), "$options": "i"}
        query["$or"] = [{"invoice_no": rx}, {"po_number": rx}, {"asn_number": rx}, {"dispatch_no": rx}, {"transporter": rx}]
    page = max(1, page)
    page_size = min(max(1, page_size), 200)
    total = await db.asn_creation.count_documents(query)
    docs = await db.asn_creation.find(query).sort("created_at", -1).skip((page - 1) * page_size).to_list(page_size)
    return {"items": [serialize(d) for d in docs], "total": total, "page": page,
            "pages": max(1, -(-total // page_size))}


@router.put("/records/{record_id}")
async def edit_asn(record_id: str, body: AsnEditInput, user: dict = Depends(get_current_user)):
    if not ObjectId.is_valid(record_id):
        raise HTTPException(status_code=400, detail="Invalid ID")
    doc = await db.asn_creation.find_one({"_id": ObjectId(record_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="ASN record not found")
    if doc["status"] in ("Processing", "Completed"):
        raise HTTPException(status_code=400, detail=f"Cannot edit a {doc['status']} record")
    update = {"po_number": body.po_number.strip(), "transporter": body.transporter.strip(),
              "basic_amount": body.basic_amount, "total_amount": body.total_amount, "updated_at": now_iso()}
    merged = {**doc, **update}
    update["status"] = compute_status(merged)
    await db.asn_creation.update_one({"_id": doc["_id"]}, {"$set": update})
    if body.po_number.strip() and ObjectId.is_valid(doc.get("master_dispatch_id", "")):
        await db.master_dispatch.update_one({"_id": ObjectId(doc["master_dispatch_id"])},
                                            {"$set": {"po_number": body.po_number.strip(), "updated_at": now_iso()}})
    await log_activity(user["username"], "asn_edited", f"{doc.get('invoice_no')} PO {body.po_number}", "asn")
    return serialize(await db.asn_creation.find_one({"_id": doc["_id"]}))


@router.post("/records/{record_id}/pdi")
async def upload_pdi(record_id: str, file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    if not ObjectId.is_valid(record_id):
        raise HTTPException(status_code=400, detail="Invalid ID")
    doc = await db.asn_creation.find_one({"_id": ObjectId(record_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="ASN record not found")
    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="PDI must be a PDF file")
    content = await file.read()
    if not content.startswith(b"%PDF"):
        raise HTTPException(status_code=400, detail="Invalid PDF file")
    fid = uuid.uuid4().hex
    path = PDI_DIR / f"{fid}.pdf"
    path.write_bytes(content)
    update = {"pdi_file_path": str(path), "pdi_file_name": file.filename, "updated_at": now_iso()}
    if doc["status"] in ("Draft", "Ready"):
        update["status"] = compute_status({**doc, **update})
    await db.asn_creation.update_one({"_id": doc["_id"]}, {"$set": update})
    await log_activity(user["username"], "asn_pdi_uploaded", f"{doc.get('invoice_no')}: {file.filename}", "asn")
    return {"ok": True, "pdi_file_name": file.filename}


@router.get("/stats")
async def asn_stats(user: dict = Depends(get_current_user)):
    today = datetime.now(timezone.utc).date().isoformat()
    return {
        "total": await db.asn_creation.count_documents({}),
        "ready": await db.asn_creation.count_documents({"status": "Ready"}),
        "draft": await db.asn_creation.count_documents({"status": "Draft"}),
        "processing": await db.asn_creation.count_documents({"status": "Processing"}),
        "completed": await db.asn_creation.count_documents({"status": "Completed"}),
        "failed": await db.asn_creation.count_documents({"status": "Failed"}),
        "today": await db.asn_creation.count_documents({"status": "Completed", "completed_at": {"$gte": today}}),
    }


# ---------- Runs (queue: one ASN at a time) ----------

async def _resolve_documents(ids: list[str]):
    """Auto-attach required documents from Master Dispatch and split runnable vs blocked ids."""
    required = await db.document_types.find({"required_for_asn": True, "active": True}).to_list(50)
    runnable, blocked = [], []
    for rid in ids:
        doc = await db.asn_creation.find_one({"_id": ObjectId(rid)})
        if not doc:
            continue
        md = None
        if ObjectId.is_valid(doc.get("master_dispatch_id", "")):
            md = await db.master_dispatch.find_one({"_id": ObjectId(doc["master_dispatch_id"])},
                                                   {"documents": 1})
        docs_map = {d.get("type"): d for d in (md or {}).get("documents") or []}
        pdi_doc = docs_map.get("PDI")
        if pdi_doc and pdi_doc.get("file_path") and os.path.exists(pdi_doc["file_path"]) \
                and doc.get("pdi_file_path") != pdi_doc["file_path"]:
            await db.asn_creation.update_one({"_id": doc["_id"]}, {"$set": {
                "pdi_file_path": pdi_doc["file_path"],
                "pdi_file_name": f"{pdi_doc.get('report_no', 'PDI')}.pdf", "updated_at": now_iso()}})
            doc["pdi_file_path"] = pdi_doc["file_path"]
        missing = []
        for t in required:
            entry = docs_map.get(t["key"])
            has_file = bool(entry and entry.get("file_path") and os.path.exists(entry["file_path"]))
            if t["key"] == "PDI" and not has_file:
                has_file = bool(doc.get("pdi_file_path") and os.path.exists(doc["pdi_file_path"]))
            if not has_file:
                missing.append(t["label"])
        if missing:
            msg = f"Required document(s) missing: {', '.join(missing)}. Generate the PDI in the AI PDI Generator — it attaches to the dispatch automatically."
            await db.asn_creation.update_one({"_id": doc["_id"]}, {"$set": {
                "status": "Failed", "error_message": msg, "updated_at": now_iso()}})
            blocked.append({"invoice_no": doc.get("invoice_no", ""), "missing": missing})
        else:
            runnable.append(rid)
    return runnable, blocked


async def _start(background_tasks: BackgroundTasks, ids: list[str], user: str):
    await get_mode()  # blocks in maintenance / emergency stop before scheduling
    if run_state["running"]:
        raise HTTPException(status_code=409, detail="An ASN automation run is already in progress")
    if not ids:
        raise HTTPException(status_code=400, detail="No records to process")
    runnable, blocked = await _resolve_documents(ids)
    if not runnable:
        detail = "; ".join(f"{b['invoice_no']}: missing {', '.join(b['missing'])}" for b in blocked) or "No valid records"
        raise HTTPException(status_code=400, detail=f"ASN blocked — {detail}")

    if desktop_execution_enabled():
        mode = await get_mode()
        is_test = mode != "live"
        await require_desktop_worker("asn_creation", allow_offline_test=True, test_mode=is_test)
        jobs = []
        for rec_id in runnable:
            doc = await db.asn_creation.find_one({"_id": ObjectId(rec_id)})
            if not doc:
                continue
            payload = {
                "po_number": doc.get("po_number", ""),
                "invoice_no": doc.get("invoice_no", ""),
                "invoice_date": to_dmy(doc.get("invoice_date", "")),
                "basic_amount": doc.get("basic_amount", 0),
                "total_amount": doc.get("total_amount", 0),
                "cgst": doc.get("cgst", 0), "sgst": doc.get("sgst", 0),
                "igst": doc.get("igst", 0), "no_of_cases": doc.get("no_of_cases", 0),
                "transporter": doc.get("transporter", ""), "items": doc.get("items", []),
                "pdi_path": doc.get("pdi_file_path", ""),
                "document_key": "pdi",
                "batch_allocations": {
                    allocation.get("part_number"): [{
                        "batch_no": allocation.get("batch_number"),
                        "allocate_qty": allocation.get("allocated_quantity", 0),
                        "consider": allocation.get("batch_considerable") == "Yes",
                    }]
                    for allocation in (doc.get("batch_allocations") or [])
                    if allocation.get("part_number") and allocation.get("batch_number")
                },
            }
            job = await create_automation_job(
                job_type="asn_creation", payload=payload, source_record_id=rec_id,
                created_by=user, test_mode=is_test, priority=50,
            )
            await db.asn_creation.update_one(
                {"_id": ObjectId(rec_id)},
                {"$set": {"status": "Queued", "desktop_job_id": job["id"],
                          "error_message": "", "updated_at": now_iso()}},
            )
            jobs.append(job)
        return {"execution": "desktop", "queued": len(jobs), "jobs": jobs, "skipped": blocked}

    run_id = str(uuid.uuid4())
    run_state.update({"running": True, "run_id": run_id, "total": len(runnable), "processed": 0,
                      "current": None, "started_at": now_iso()})
    background_tasks.add_task(process_queue, runnable, run_id, user)
    return {"run_id": run_id, "total": len(runnable), "skipped": blocked}


@router.post("/run")
async def asn_run(req: RunRequest, background_tasks: BackgroundTasks, user: dict = Depends(get_current_user)):
    ids = [i for i in req.ids if ObjectId.is_valid(i)]
    result = await _start(background_tasks, ids, user["username"])
    await log_activity(user["username"], "asn_run", f"{len(ids)} record(s)", "asn")
    return result


@router.post("/run-ready")
async def asn_run_ready(background_tasks: BackgroundTasks, user: dict = Depends(get_current_user)):
    ids = [str(d["_id"]) async for d in db.asn_creation.find({"status": "Ready", **(await env_list_filter())}, {"_id": 1})]
    result = await _start(background_tasks, ids, user["username"])
    await log_activity(user["username"], "asn_run_ready", f"{len(ids)} record(s)", "asn")
    return result


@router.post("/retry-failed")
async def asn_retry_failed(background_tasks: BackgroundTasks, user: dict = Depends(get_current_user)):
    ids = [str(d["_id"]) async for d in db.asn_creation.find({"status": "Failed", **(await env_list_filter())}, {"_id": 1})]
    result = await _start(background_tasks, ids, user["username"])
    await log_activity(user["username"], "asn_retry_failed", f"{len(ids)} record(s)", "asn")
    return result


@router.get("/run-status")
async def asn_run_status(user: dict = Depends(get_current_user)):
    return run_state


# ---------- Batch allocation ----------

class AllocationRow(BaseModel):
    batch_no: str
    allocate_qty: float = Field(0, ge=0)
    consider: bool = True


class AllocationConfirm(BaseModel):
    record_id: str
    allocations: list[AllocationRow]


@router.post("/allocation/confirm")
async def confirm_allocation(body: AllocationConfirm, user: dict = Depends(get_current_user)):
    if not alloc_state.get("event") or alloc_state.get("record_id") != body.record_id:
        raise HTTPException(status_code=409, detail="No batch allocation is currently awaiting for this record")
    by_no = {b["batch_no"]: b for b in alloc_state["batches"]}
    total = 0.0
    for a in body.allocations:
        b = by_no.get(a.batch_no)
        if not b:
            raise HTTPException(status_code=400, detail=f"Unknown batch {a.batch_no}")
        if a.allocate_qty > float(b["available_qty"]) + 0.001:
            raise HTTPException(status_code=400, detail="Allocation cannot exceed Available Quantity.")
        total += a.allocate_qty
    if abs(total - float(alloc_state["asn_qty"])) > 0.001:
        raise HTTPException(status_code=400,
                            detail=f"Total Allocated ({total:g}) must equal ASN Quantity ({alloc_state['asn_qty']:g})")
    alloc_state["result"] = [a.model_dump() for a in body.allocations]
    alloc_state["cancelled"] = False
    alloc_state["event"].set()
    await log_activity(user["username"], "asn_batch_allocated", f"{len(body.allocations)} batch(es) confirmed", "asn")
    return {"ok": True}


@router.post("/allocation/cancel")
async def cancel_allocation(body: dict, user: dict = Depends(get_current_user)):
    if not alloc_state.get("event") or alloc_state.get("record_id") != body.get("record_id"):
        raise HTTPException(status_code=409, detail="No batch allocation is currently awaiting for this record")
    alloc_state["cancelled"] = True
    alloc_state["result"] = None
    alloc_state["event"].set()
    await log_activity(user["username"], "asn_batch_allocation_cancelled", "", "asn")
    return {"ok": True}


@router.get("/batch-allocations")
async def batch_allocations(search: Optional[str] = None, user: dict = Depends(get_current_user)):
    query = {}
    if search:
        rx = {"$regex": re.escape(search), "$options": "i"}
        query["$or"] = [{"asn_number": rx}, {"invoice_no": rx}, {"part_number": rx},
                        {"batch_number": rx}, {"dispatch_no": rx}]
    docs = await db.asn_batch_allocations.find(query).sort("created_at", -1).to_list(500)
    return {"items": [serialize(d) for d in docs], "total": len(docs)}


@router.get("/export")
async def asn_export(user: dict = Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    docs = await db.asn_creation.find(await env_list_filter()).sort("created_at", -1).to_list(5000)
    wb = Workbook()
    ws = wb.active
    ws.title = "ASN Creation"
    ws.append(["Dispatch No", "Invoice No", "Invoice Date", "PO Number", "Transporter", "Plant",
               "Basic Amount", "Total Amount", "Parts", "PDI File", "Status", "ASN Number", "Batch Allocations", "Error"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    for d in docs:
        allocs = "; ".join(f"{a['part_number']}: {a['batch_number']}={a['allocated_quantity']:g} ({a['batch_considerable']})"
                           for a in (d.get("batch_allocations") or []))
        ws.append([d.get("dispatch_no"), d.get("invoice_no"), d.get("invoice_date"), d.get("po_number"),
                   d.get("transporter"), d.get("plant"), d.get("basic_amount"), d.get("total_amount"),
                   ", ".join(i["part_number"] for i in d.get("items", [])), d.get("pdi_file_name", ""),
                   d.get("status"), d.get("asn_number"), allocs, d.get("error_message")])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=asn_creation.xlsx"})


@router.get("/screenshots/{name}")
async def asn_screenshot(name: str, user: dict = Depends(get_current_user)):
    if "/" in name or ".." in name:
        raise HTTPException(status_code=400, detail="Invalid name")
    path = ROOT_DIR / "screenshots" / name
    if not path.exists():
        raise HTTPException(status_code=404, detail="Not found")
    media = "text/html" if name.endswith(".html") else "image/png"
    return FileResponse(str(path), media_type=media)
